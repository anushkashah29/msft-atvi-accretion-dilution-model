"""
generate_model.py
Generates MSFT_ATVI_Accretion_Dilution_Model.xlsx from scratch.
Run:  python3 generate_model.py
Output: MSFT_ATVI_Accretion_Dilution_Model.xlsx (same directory)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY        = "FF1B3A6B"   # title bars, primary brand
MED_BLUE    = "FF2A4F8A"   # column-header bars
WHITE       = "FFFFFFFF"
GOLD        = "FFC9A84C"   # subtitle text
GRAY_BODY   = "FF5A6070"   # regular body text
LT_GRAY     = "FFF4F5F7"   # section-header / subtotal rows
LT_BLUE     = "FFE6F1FB"   # key figures (NI, EPS)
LT_GREEN    = "FFEAF3DE"   # synergy / positive items
LT_RED      = "FFFCEBEB"   # cost / drag items
LT_AMBER    = "FFFAEEDA"   # amortisation totals, accretion row
LT_MINT     = "FFEBF5EE"   # interview-insight callout
DK_RED      = "FF791F1F"   # after-tax financing cost text
DK_GREEN    = "FF085041"   # after-tax synergy text
DK_AMBER    = "FF633806"   # amortisation text
GREEN_CELL  = "FF008000"   # ATVI / synergy numbers in pro forma
RED_CELL    = "FFFF0000"   # drag numbers in pro forma
BLACK       = "FF000000"
DARK_NAVY   = "FF1B3A6B"   # reused alias

# Tab colours
TAB_GOLD    = "C9A84C"
TAB_BLUE    = "185FA5"
TAB_AMBER   = "FAEEDA"
TAB_RED     = "993C1D"
TAB_GREEN   = "3B6D11"
TAB_NAVY    = "1B3A6B"

# ── Number formats ────────────────────────────────────────────────────────────
NF_DOLLAR   = r'$#,##0;"($"#,##0)\;-'
NF_DOLLAR2  = r'$#,##0.00;"($"#,##0.00)\;-'
NF_PCT      = r'0.0%;\(0.0%\)\;-'
NF_MULT     = r'0.0\x;\(0.0"x)"\;-'
NF_GENERAL  = "General"

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color=BLACK, bold=False, size=10, italic=False, name="Calibri"):
    return Font(name=name, color=hex_color, bold=bold, size=size, italic=italic)

def align(horiz="left", wrap=False, vert="center"):
    return Alignment(horizontal=horiz, vertical=vert, wrap_text=wrap)

def style(ws, coord, value=None, fg=None, fc=BLACK, bold=False, size=10,
          horiz="left", wrap=False, nf=NF_GENERAL, italic=False):
    cell = ws[coord]
    if value is not None:
        cell.value = value
    if fg:
        cell.fill = fill(fg)
    cell.font = font(fc, bold, size, italic)
    cell.alignment = align(horiz, wrap)
    if nf != NF_GENERAL:
        cell.number_format = nf
    return cell

def title_row(ws, row, text, subtitle=False):
    cell = ws[f"B{row}"]
    cell.value = text
    cell.fill = fill(NAVY)
    cell.font = font(GOLD if subtitle else WHITE, bold=not subtitle,
                     size=10 if subtitle else 13)
    cell.alignment = align("left")

def col_header(ws, coord, text, horiz="right"):
    cell = ws[coord]
    cell.value = text
    cell.fill = fill(MED_BLUE)
    cell.font = font(WHITE, bold=True, size=10)
    cell.alignment = align(horiz)

def section_hdr(ws, row, text, col="B", fg=LT_GRAY, fc=DARK_NAVY):
    cell = ws[f"{col}{row}"]
    cell.value = text
    cell.fill = fill(fg)
    cell.font = font(fc, bold=True, size=10)
    cell.alignment = align("left")

def data_cell(ws, coord, value, fg=None, fc=BLACK, bold=False,
              horiz="right", nf=NF_DOLLAR, italic=False):
    cell = ws[coord]
    cell.value = value
    if fg:
        cell.fill = fill(fg)
    cell.font = font(fc, bold, italic=italic)
    cell.alignment = align(horiz)
    cell.number_format = nf
    return cell

def label(ws, coord, text, fg=None, fc=BLACK, bold=False, wrap=False, italic=False):
    cell = ws[coord]
    cell.value = text
    if fg:
        cell.fill = fill(fg)
    cell.font = font(fc, bold, italic=italic)
    cell.alignment = align("left", wrap)

def note(ws, coord, text, fg=WHITE, fc=GRAY_BODY, italic=False):
    cell = ws[coord]
    cell.value = text
    cell.fill = fill(fg)
    cell.font = font(fc, italic=italic)
    cell.alignment = align("left", wrap=True)

# ── Workbook scaffold ─────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)   # remove default blank sheet

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 0 — How To: Step by Step
# ══════════════════════════════════════════════════════════════════════════════
ws0 = wb.create_sheet("How To — Step by Step")
ws0.sheet_properties.tabColor = TAB_GOLD
ws0.column_dimensions["A"].width = 2
ws0.column_dimensions["B"].width = 38
ws0.column_dimensions["C"].width = 18
ws0.column_dimensions["D"].width = 22
ws0.column_dimensions["E"].width = 18
ws0.column_dimensions["F"].width = 22

title_row(ws0, 2, "M&A ACCRETION / DILUTION MODEL — STEP-BY-STEP GUIDE")
title_row(ws0, 3, "Microsoft / Activision Blizzard  |  Deal closed October 13, 2023  |  $68.7B all-cash", subtitle=True)

rows_howto = [
    (5,  "What is accretion / dilution analysis?", LT_GRAY, DARK_NAVY, True),
    (6,  ("Accretion / dilution measures whether an acquisition INCREASES or DECREASES the acquirer's "
          "earnings per share (EPS). If pro forma EPS > standalone EPS, the deal is ACCRETIVE. If pro "
          "forma EPS < standalone EPS, it is DILUTIVE. Every IB analyst builds this model for every "
          "deal. It is the primary financial health check before a board approves a transaction."),
         None, GRAY_BODY, False),
    (8,  "  STEP 1: Sources & Uses of Funds", NAVY, WHITE, True),
    (9,  "What you do:", None, DARK_NAVY, True),
    (10, ("Determine how the acquisition is financed — existing cash, new debt, or new equity (stock). "
          "For Microsoft/Activision: 100% cash deal. $20B from MSFT's balance sheet + $50.5B in new "
          "debt. No new shares issued → share count unchanged → EPS denominator is fixed."),
         None, GRAY_BODY, False),
    (11, "→ See tab: '1 - Sources & Uses'", None, DARK_NAVY, False),
    (12, "Why it matters:", None, DARK_NAVY, True),
    (13, ("Why it matters: the financing mix determines the interest cost (Step 4) and whether the share "
          "count changes. All-cash deals are simpler — only the numerator (net income) changes."),
         None, GRAY_BODY, False),
    (15, "  STEP 2: Standalone Financials", NAVY, WHITE, True),
    (16, "What you do:", None, DARK_NAVY, True),
    (17, ("Collect the acquirer's (MSFT) and target's (ATVI) most recent income statements. Calculate "
          "standalone EPS for MSFT — this is your baseline. Calculate ATVI's net income — this is what "
          "you're 'buying.'"), None, GRAY_BODY, False),
    (18, "→ See tab: '2 - Standalone P&L'", None, DARK_NAVY, False),
    (19, "Why it matters:", None, DARK_NAVY, True),
    (20, ("Why it matters: you need to know MSFT's current EPS ($10.29) and ATVI's net income ($3,572M) "
          "before you can calculate whether the deal adds or destroys per-share value."),
         None, GRAY_BODY, False),
    (22, "  STEP 3: Purchase Price Allocation (PPA)", NAVY, WHITE, True),
    (23, "What you do:", None, DARK_NAVY, True),
    (24, ("When you pay more than book value for a company, GAAP requires you to allocate the premium to "
          "identifiable intangible assets (game titles, customer relationships, brand). These intangibles "
          "are then AMORTISED through the income statement — creating a real EPS headwind every year "
          "post-close. Goodwill (~$46B) is NOT amortised but IS tested annually for impairment."),
         None, GRAY_BODY, False),
    (25, "→ See tab: '3 - PPA & Amortisation'", None, DARK_NAVY, False),
    (26, "Why it matters:", None, DARK_NAVY, True),
    (27, ("Why it matters: this is the most commonly missed step. The ~$1.59B annual after-tax "
          "amortisation alone costs ~$0.21 EPS per year."), None, GRAY_BODY, False),
    (29, "  STEP 4: Financing Cost", NAVY, WHITE, True),
    (30, "What you do:", None, DARK_NAVY, True),
    (31, ("Calculate the annual interest expense on the new debt raised to fund the acquisition. Also "
          "account for the FOREGONE interest income on cash deployed (opportunity cost). Both are "
          "tax-deductible, so only the after-tax cost hits EPS."), None, GRAY_BODY, False),
    (32, "→ See tab: '4 - Financing Cost'", None, DARK_NAVY, False),
    (33, "Why it matters:", None, DARK_NAVY, True),
    (34, ("Why it matters: $50.5B of new debt at ~4.6% costs $2.3B/yr in gross interest plus $900M "
          "foregone income. After tax (18.8%), that's ~$2.6B — or ~$0.35 EPS drag."),
         None, GRAY_BODY, False),
    (36, "  STEP 5: Synergies", NAVY, WHITE, True),
    (37, "What you do:", None, DARK_NAVY, True),
    (38, ("Estimate the annual cost savings and revenue uplift from combining the two companies. Apply a "
          "REALISATION SCHEDULE (40% Year 1, 70% Year 2, 100% Year 3). Synergies are tax-effected. "
          "COST synergies are more credible — discount revenue synergies by 50% in a conservative case."),
         None, GRAY_BODY, False),
    (39, "→ See tab: '5 - Synergies'", None, DARK_NAVY, False),
    (41, "  STEP 6: Pro Forma EPS — The Verdict", NAVY, WHITE, True),
    (42, "What you do:", None, DARK_NAVY, True),
    (43, ("Combine all five steps: MSFT standalone NI + ATVI NI − after-tax financing cost − after-tax "
          "PPA amortisation + after-tax synergies = Pro Forma NI. Divide by the (unchanged) share count "
          "to get Pro Forma EPS. Compare against standalone EPS."), None, GRAY_BODY, False),
    (44, "→ See tab: '6 - Pro Forma EPS'", None, DARK_NAVY, False),
    (45, "Why it matters:", None, DARK_NAVY, True),
    (46, ("This is the final answer every board and analyst cares about. Is the deal accretive or "
          "dilutive? By how much? How sensitive is the verdict to synergy assumptions?"),
         None, GRAY_BODY, False),
]

for r, txt, fg, fc, bold in rows_howto:
    cell = ws0[f"B{r}"]
    cell.value = txt
    if fg:
        cell.fill = fill(fg)
    cell.font = font(fc, bold)
    cell.alignment = align("left", wrap=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Sources & Uses
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("1 - Sources & Uses")
ws1.sheet_properties.tabColor = TAB_BLUE
ws1.column_dimensions["A"].width = 2
ws1.column_dimensions["B"].width = 36
ws1.column_dimensions["C"].width = 16
ws1.column_dimensions["D"].width = 14
ws1.column_dimensions["E"].width = 42

title_row(ws1, 2, "STEP 1: SOURCES & USES OF FUNDS")
title_row(ws1, 3, "How is the $68.7B acquisition financed?  |  All-cash deal  |  No new shares issued",
          subtitle=True)

col_header(ws1, "B5", "Line Item", horiz="left")
col_header(ws1, "C5", "Amount ($M)")
col_header(ws1, "D5", "% of Total")
col_header(ws1, "E5", "Notes / Source")

# USES section
section_hdr(ws1, 6, "  USES OF FUNDS")

uses = [
    (7,  "  Equity purchase price (equity value)",    68700,
         "MSFT press release Jan 18 2022. $95.00/share × 723.7M ATVI diluted shares"),
    (8,  "  Transaction fees (advisory, legal, financing)", 800,
         "Estimated; typical 1.0–1.5% of deal value for complex cross-border transactions"),
    (9,  "  Refinancing of ATVI net debt assumed",    970,
         "ATVI LTM net debt position at close; must be repaid or refinanced on acquisition"),
]
for r, lbl, amt, nt in uses:
    label(ws1, f"B{r}", lbl)
    data_cell(ws1, f"C{r}", amt, nf=NF_DOLLAR)
    note(ws1, f"E{r}", nt)

# TOTAL USES row
for col, v, nf in [("B", "  TOTAL USES", NF_GENERAL),
                   ("C", "=SUM(C7:C9)", NF_DOLLAR),
                   ("D", "=C10/C10", NF_PCT)]:
    cell = ws1[f"{col}10"]
    cell.value = v
    cell.fill = fill(LT_GRAY)
    cell.font = font(DARK_NAVY if col == "B" else BLACK, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col != "B":
        cell.number_format = nf

# SOURCES section
section_hdr(ws1, 12, "  SOURCES OF FUNDS")

sources = [
    (13, "  Cash from MSFT balance sheet",            20000,
         "MSFT had ~$111B liquid assets; deployed $20B to minimise new debt quantum"),
    (14, "  New long-term bonds issued (10–30 year)", 35000,
         "MSFT Aaa/AAA rated — issued at historically tight spreads given credit quality"),
    (15, "  New medium-term notes (3–5 year)",        10000,
         "Diversified maturity profile to reduce refinancing risk concentration"),
    (16, "  New commercial paper / short-term",        5470,
         "Bridged until long-term bonds fully placed in market"),
]
for r, lbl, amt, nt in sources:
    label(ws1, f"B{r}", lbl)
    data_cell(ws1, f"C{r}", amt, nf=NF_DOLLAR)
    note(ws1, f"E{r}", nt)

# TOTAL SOURCES row
for col, v, nf in [("B", "  TOTAL SOURCES", NF_GENERAL),
                   ("C", "=SUM(C13:C16)", NF_DOLLAR),
                   ("D", "=C18/C18", NF_PCT)]:
    cell = ws1[f"{col}18"]
    cell.value = v
    cell.fill = fill(LT_GRAY)
    cell.font = font(DARK_NAVY if col == "B" else BLACK, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col != "B":
        cell.number_format = nf

# CHECK row
label(ws1, "B20", "CHECK: Sources − Uses (should be $0)")
data_cell(ws1, "C20", "=C18-C10", nf=NF_DOLLAR)
note(ws1, "E20", "Must equal zero — if not, check for input errors in Sources or Uses above")

# FINANCING MIX section
section_hdr(ws1, 22, "  FINANCING MIX BREAKDOWN")

mix = [
    (23, "  Cash from balance sheet", "=C13"),
    (24, "  Long-term bonds",         "=C14"),
    (25, "  Medium-term notes",       "=C15"),
    (26, "  Commercial paper",        "=C16"),
]
for r, lbl, val in mix:
    label(ws1, f"B{r}", lbl)
    data_cell(ws1, f"C{r}", val, nf=NF_DOLLAR)
    data_cell(ws1, f"D{r}", f"={get_column_letter(3)}{r}/C18", nf=NF_PCT)

# Key implication callout
cell = ws1["B28"]
cell.value = ("Key implication: NO new MSFT shares issued — All-cash deal → share count unchanged "
              "→ EPS denominator fixed at 7,434M diluted shares throughout the model")
cell.fill = fill(LT_MINT)
cell.font = font(DK_GREEN, bold=True)
cell.alignment = align("left", wrap=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Standalone P&L
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2 - Standalone P&L")
ws2.sheet_properties.tabColor = TAB_BLUE
ws2.column_dimensions["A"].width = 2
ws2.column_dimensions["B"].width = 38
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 42

title_row(ws2, 2, "STEP 2: STANDALONE INCOME STATEMENTS")
title_row(ws2, 3, "Pre-deal income statements  |  FY2023E at time of close  |  $M", subtitle=True)

col_header(ws2, "B5", "Line Item", horiz="left")
col_header(ws2, "C5", "Microsoft (MSFT)")
col_header(ws2, "D5", "Activision (ATVI)")
col_header(ws2, "E5", "Notes")

# Raw inputs
pl_inputs = [
    (7,  "  Revenue",                    211915,   8803,
         "MSFT: FY2023 10-K. ATVI: LTM at close per last filed 10-Q"),
    (8,  "  Cost of Revenue",            -65863,  -2892,
         "Direct costs of products/services sold"),
    (10, "  Research & Development",     -29510,  -1040,
         "MSFT R&D; ATVI game development costs"),
    (11, "  Sales, General & Administrative", -24456, -1101,
         "Corporate overhead, marketing, admin"),
    (13, "  Interest Income / (Expense), net",  1413,   218,
         "MSFT net interest on $111B liquid assets"),
    (14, "  Other Income / (Expense)",      -78,   -35,
         "FX, equity method investments"),
    (16, "  Income Tax Expense",          -16950,  -381,
         "MSFT effective rate 18.8%; ATVI ~20.0%"),
]
for r, lbl, msft, atvi, nt in pl_inputs:
    label(ws2, f"B{r}", lbl)
    data_cell(ws2, f"C{r}", msft, nf=NF_DOLLAR)
    data_cell(ws2, f"D{r}", atvi, nf=NF_DOLLAR)
    note(ws2, f"E{r}", nt)

# Subtotal rows (light gray)
for row, lbl, cf, df in [
    (9,  "  Gross Profit",            "=C7+C8",        "=D7+D8"),
    (12, "  Operating Income (EBIT)", "=C9+C10+C11",   "=D9+D10+D11"),
    (15, "  Pre-Tax Income (EBT)",    "=C12+C13+C14",  "=D12+D13+D14"),
]:
    for col, val in [("B", lbl), ("C", cf), ("D", df)]:
        cell = ws2[f"{col}{row}"]
        cell.value = val
        cell.fill = fill(LT_GRAY)
        cell.font = font(DARK_NAVY if col == "B" else BLACK, bold=True)
        cell.alignment = align("left" if col == "B" else "right")
        if col in ("C", "D"):
            cell.number_format = NF_DOLLAR

# Net Income (highlighted in light blue)
for col, val, fc_c, bold_c in [
    ("B", "  Net Income",  DARK_NAVY, True),
    ("C", "=C15+C16",      DARK_NAVY, True),
    ("D", "=D15+D16",      DARK_NAVY, True),
]:
    cell = ws2[f"{col}17"]
    cell.value = val
    cell.fill = fill(LT_BLUE)
    cell.font = font(fc_c, bold_c)
    cell.alignment = align("left" if col == "B" else "right")
    if col in ("C", "D"):
        cell.number_format = NF_DOLLAR

# Diluted shares
label(ws2, "B19", "  Diluted Shares Outstanding (M)")
data_cell(ws2, "C19", 7434, nf=NF_DOLLAR)
data_cell(ws2, "D19", 724,  nf=NF_DOLLAR)
note(ws2, "E19", "MSFT: 10-K. ATVI: pre-deal shares before cash settlement")

# EPS (highlighted)
for col, val, fc_c in [
    ("B", "  Diluted EPS ($)", DARK_NAVY),
    ("C", "=C17/C19",          DARK_NAVY),
    ("D", "=D17/D19",          DARK_NAVY),
]:
    cell = ws2[f"{col}20"]
    cell.value = val
    cell.fill = fill(LT_BLUE)
    cell.font = font(fc_c, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col in ("C", "D"):
        cell.number_format = NF_DOLLAR2

# Effective tax rate (gray subtotal)
for col, val in [("B", "  Effective Tax Rate"), ("C", "=ABS(C16)/C15"), ("D", "=ABS(D16)/D15")]:
    cell = ws2[f"{col}21"]
    cell.value = val
    cell.fill = fill(LT_GRAY)
    cell.font = font(DARK_NAVY if col == "B" else BLACK, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col in ("C", "D"):
        cell.number_format = NF_PCT

# P/E multiple
label(ws2, "B23", "  P/E Multiple (at deal price / current price)")
data_cell(ws2, "C23", 409, nf=NF_GENERAL)
data_cell(ws2, "D23", 95,  nf=NF_GENERAL)

# EBITDA section
label(ws2, "B25", "  EBITDA")
data_cell(ws2, "C25", 105672, nf=NF_DOLLAR)
data_cell(ws2, "D25", 1820,   nf=NF_DOLLAR)
note(ws2, "E25", "EBITDA = EBIT + D&A; used for EV/EBITDA valuation multiple")

label(ws2, "B26", "  EV/EBITDA multiple")
for col, val in [("C", "=C25/C24"), ("D", "=D25/D24")]:
    cell = ws2[f"{col}26"]
    cell.value = val
    cell.fill = fill(LT_GRAY)
    cell.font = font(BLACK, bold=True)
    cell.alignment = align("right")
    cell.number_format = NF_MULT

# EV proxies (unlabelled inputs)
label(ws2, "B24", "  EV/EBITDA (at deal price / market cap)")
data_cell(ws2, "C24", 2452000, nf=NF_DOLLAR)
data_cell(ws2, "D24", 68700,   nf=NF_DOLLAR)

ws2["B28"].value = ("KEY: MSFT standalone EPS (row 20, column C) is the BASELINE — "
                    "every pro forma calculation compares against this number.")
ws2["B28"].font = font(DARK_NAVY, bold=True)
ws2["B28"].alignment = align("left", wrap=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — PPA & Amortisation
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3 - PPA & Amortisation")
ws3.sheet_properties.tabColor = TAB_AMBER
ws3.column_dimensions["A"].width = 2
ws3.column_dimensions["B"].width = 36
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 13
ws3.column_dimensions["F"].width = 13
ws3.column_dimensions["G"].width = 14
ws3.column_dimensions["H"].width = 42

title_row(ws3, 2, "STEP 3: PURCHASE PRICE ALLOCATION (PPA) & AMORTISATION")
title_row(ws3, 3, "Intangible asset step-up and annual amortisation schedule  |  $M", subtitle=True)

for coord, hdr, h in [("B5","Intangible Asset","left"),("C5","Fair Value ($M)","right"),
                       ("D5","Useful Life (yrs)","right"),("E5","Annual Amort.","right"),
                       ("F5","Tax Deductible?","right"),("G5","After-Tax Amort.","right"),
                       ("H5","Notes","right")]:
    col_header(ws3, coord, hdr, horiz=h)

section_hdr(ws3, 7, "  IDENTIFIABLE INTANGIBLE ASSETS ACQUIRED")

# Intangible assets
intangibles = [
    (8,  "  Developed game titles (CoD, WoW, Overwatch, etc.)", 8500,  6, "Yes",
         "=IF(D8>0,C8/D8,0)",
         "=IF(F8=\"Yes\",E8*(1-'4 - Financing Cost'!$C$12),E8)",
         "Core IP; amortised over estimated economic life of franchise"),
    (9,  "  Customer relationships / active player base",        3200, 10, "Yes",
         "=IF(D9>0,C9/D9,0)",
         "=IF(F9=\"Yes\",E9*(1-'4 - Financing Cost'!$C$12),E9)",
         "300M+ monthly active users; relationship value amortised over retention curve"),
    (10, "  In-process R&D (game pipeline)",                     1400,  0, "No",
         0, 0,
         "Expensed immediately at acquisition close per ASC 805; no ongoing amortisation"),
    (11, "  Technology platform / engine",                       1100,  5, "Yes",
         "=IF(D11>0,C11/D11,0)",
         "=IF(F11=\"Yes\",E11*(1-'4 - Financing Cost'!$C$12),E11)",
         "Proprietary game development tools and engine technology"),
    (12, "  Trade names and brand (Blizzard, King, Activision)", 1800,  0, "No",
         0, 0,
         "Indefinite-life intangible — NOT amortised, tested annually for impairment"),
]
for r, lbl, fv, life, td, amort, at_amort, nt in intangibles:
    label(ws3, f"B{r}", lbl)
    data_cell(ws3, f"C{r}", fv, nf=NF_DOLLAR)
    ws3[f"D{r}"].value = life
    ws3[f"D{r}"].alignment = align("right")
    ws3[f"F{r}"].value = td
    ws3[f"F{r}"].alignment = align("right")
    for col, val in [("E", amort), ("G", at_amort)]:
        cell = ws3[f"{col}{r}"]
        cell.value = val
        cell.alignment = align("right")
        cell.number_format = NF_DOLLAR

    note(ws3, f"H{r}", nt)

# Goodwill row (uses fixed ATVI net tangible book value ~$8,470M)
label(ws3, "B14", "  Goodwill (residual premium)")
ws3["C14"].value = "='1 - Sources & Uses'!C10-SUM(C8:C12)-8470"
ws3["C14"].alignment = align("right")
ws3["C14"].number_format = NF_DOLLAR
ws3["D14"].value = "N/A"
ws3["E14"].value = "-"
ws3["F14"].value = "No"
ws3["G14"].value = "-"
note(ws3, "H14",
     "Residual = Total deal cost ($70,470M) − Identified intangibles ($16,000M) − "
     "ATVI net tangible book value ($8,470M). NOT amortised; impairment tested annually (ASC 350).")

# Total intangibles + goodwill
for col, val in [("B", "  Total Identified Intangibles + Goodwill"),
                 ("C", "=C14+SUM(C8:C12)")]:
    cell = ws3[f"{col}16"]
    cell.value = val
    cell.fill = fill(LT_GRAY)
    cell.font = font(DARK_NAVY if col == "B" else BLACK, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col == "C":
        cell.number_format = NF_DOLLAR

# ── Amortisation schedule ──────────────────────────────────────────────────
section_hdr(ws3, 18, "  ANNUAL AMORTISATION SCHEDULE (Years 1–10)")

for coord, hdr in [("B19","Asset"),("C19","Yr 1"),("D19","Yr 2"),("E19","Yr 3"),
                   ("F19","Yr 4"),("G19","Yr 5"),("H19","Yr 6–10")]:
    col_header(ws3, coord, hdr, horiz="left" if coord=="B19" else "right")

sched = [
    (20, "  Developed game titles (CoD, WoW, Overwatch, etc.)",
         ["=C8/D8"]*6, [0]),        # cols C-H: yrs 1-5 amort, yr 6+ = 0 (fully amortised)
    (21, "  Customer relationships / active player base",
         ["=C9/D9"]*6, []),
    (22, "  In-process R&D (game pipeline)",  [0]*6, []),
    (23, "  Technology platform / engine",
         ["=C11/D11"]*5 + [0], []),
    (24, "  Trade names and brand",           [0]*6, []),
]
for r, lbl, vals, _ in sched:
    label(ws3, f"B{r}", lbl)
    for i, col in enumerate("CDEFGH"):
        cell = ws3[f"{col}{r}"]
        cell.value = vals[i]
        cell.alignment = align("right")
        cell.number_format = NF_DOLLAR

# Total row (amber highlight) — corrected to sum rows 20:24
for col, val in [("B", "  Total Annual Amortisation (GAAP charge)"),
                 ("C", "=SUM(C20:C24)"), ("D", "=SUM(D20:D24)"),
                 ("E", "=SUM(E20:E24)"), ("F", "=SUM(F20:F24)"),
                 ("G", "=SUM(G20:G24)"), ("H", "=SUM(H20:H24)")]:
    cell = ws3[f"{col}26"]
    cell.value = val
    cell.fill = fill(LT_AMBER)
    cell.font = font(DK_AMBER if col == "B" else BLACK, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col != "B":
        cell.number_format = NF_DOLLAR

# ── EPS impact section ─────────────────────────────────────────────────────
section_hdr(ws3, 28, "  EPS IMPACT OF AMORTISATION (key model output)")

label(ws3, "B29", "  Annual GAAP amortisation charge ($M)")
for col, val, fg_c, fc_c in [
    ("C29", "=C26",   LT_AMBER, BLACK),
]:
    cell = ws3[col]
    cell.value = val
    cell.fill = fill(fg_c)
    cell.font = font(fc_c, bold=True)
    cell.alignment = align("right")
    cell.number_format = NF_DOLLAR
note(ws3, "H29", "Links to Pro Forma EPS model")

label(ws3, "B30", "  Tax rate (MSFT effective rate)")
ws3["C30"].value = "='4 - Financing Cost'!$C$12"   # FIXED: was C10 (interest rate)
ws3["C30"].alignment = align("right")
ws3["C30"].number_format = NF_PCT

label(ws3, "B31", "  After-tax amortisation ($M)")
for col, val in [("C31", "=C29*(1-C30)")]:
    cell = ws3[col]
    cell.value = val
    cell.fill = fill(LT_RED)
    cell.font = font(DK_RED, bold=True)
    cell.alignment = align("right")
    cell.number_format = NF_DOLLAR

label(ws3, "B32", "  EPS drag from amortisation ($/share)")
ws3["C32"].value = "=-C31/'2 - Standalone P&L'!C19"   # FIXED: removed /1000, added negative
ws3["C32"].fill = fill(LT_RED)
ws3["C32"].font = font(DK_RED, bold=True)
ws3["C32"].alignment = align("right")
ws3["C32"].number_format = NF_DOLLAR2
note(ws3, "H32", "After-tax amort ÷ MSFT diluted shares. Negative = EPS drag.")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Financing Cost
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4 - Financing Cost")
ws4.sheet_properties.tabColor = TAB_RED
ws4.column_dimensions["A"].width = 2
ws4.column_dimensions["B"].width = 38
ws4.column_dimensions["C"].width = 16
ws4.column_dimensions["D"].width = 14
ws4.column_dimensions["E"].width = 42

title_row(ws4, 2, "STEP 4: FINANCING COST — INTEREST EXPENSE ON NEW DEBT")
title_row(ws4, 3, "Annual interest cost of the $50.5B new debt raised to fund the acquisition  |  $M",
          subtitle=True)

for coord, hdr, h in [("B5","Item","left"),("C5","Amount / Rate","right"),
                       ("D5","Annual Cost ($M)","right"),("E5","Notes","right")]:
    col_header(ws4, coord, hdr, horiz=h)

section_hdr(ws4, 7, "  KEY ASSUMPTIONS")

assumptions = [
    (8,  "  Total new debt raised ($M)",            50470, None,
         "Total new debt = Long-term bonds + MTNs + Commercial paper"),
    (9,  "  Cash deployed from balance sheet ($M)", 20000, None,
         "Opportunity cost: cash was earning ~4.5% before being deployed"),
    (10, "  Blended interest rate on new debt",     0.046, NF_PCT,
         "MSFT Aaa/AAA rated. LT bonds ~4.75%, MTNs ~4.25%, CP ~5.2% — blended ~4.6%"),
    (11, "  Foregone yield on cash deployed",       0.045, NF_PCT,
         "Prior yield on MSFT liquid asset portfolio (~4.5% short-dated Treasuries)"),
    (12, "  Blended tax rate (MSFT effective rate)", 0.188, NF_PCT,
         "MSFT FY2023 effective tax rate 18.8%. Source: MSFT 10-K FY2023 Note 13"),
]
for r, lbl, val, nf_c, nt in assumptions:
    label(ws4, f"B{r}", lbl)
    cell = ws4[f"C{r}"]
    cell.value = val
    cell.alignment = align("right")
    if nf_c:
        cell.number_format = nf_c
    note(ws4, f"E{r}", nt)

section_hdr(ws4, 14, "  ANNUAL INTEREST COST BREAKDOWN")

interest_lines = [
    (15, "  Interest on new long-term bonds ($M)",        "='1 - Sources & Uses'!C14",
         "=C15*C10", "35,000 × blended LT rate"),
    (16, "  Interest on medium-term notes ($M)",          "='1 - Sources & Uses'!C15",
         "=C16*C10", None),
    (17, "  Interest on commercial paper ($M)",           "='1 - Sources & Uses'!C16",
         "=C17*C10", None),
    (18, "  Foregone interest income on cash deployed ($M)", "=C9",
         "=C9*C11",
         "Opportunity cost of using balance sheet cash. Include for conservatism."),
]
for r, lbl, c_val, d_val, nt in interest_lines:
    label(ws4, f"B{r}", lbl)
    ws4[f"C{r}"].value = c_val
    ws4[f"C{r}"].alignment = align("right")
    ws4[f"C{r}"].number_format = NF_DOLLAR
    data_cell(ws4, f"D{r}", d_val, nf=NF_DOLLAR)
    if nt:
        note(ws4, f"E{r}", nt)

# TOTAL GROSS
for col, val in [("B", "  TOTAL GROSS FINANCING COST ($M)"), ("D", "=SUM(D15:D18)")]:
    cell = ws4[f"{col}20"]
    cell.value = val
    cell.fill = fill(LT_GRAY)
    cell.font = font(DARK_NAVY if col == "B" else BLACK, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col == "D":
        cell.number_format = NF_DOLLAR

label(ws4, "B21", "  Less: tax shield on interest (deductible expense)")
data_cell(ws4, "D21", "=-D20*C12", nf=NF_DOLLAR)
note(ws4, "E21", "Interest expense is tax-deductible — reduces taxable income and therefore tax paid")

# AFTER-TAX FINANCING COST
for col, val in [("B", "  AFTER-TAX FINANCING COST ($M)"), ("D", "=D20*(1-C12)")]:
    cell = ws4[f"{col}23"]
    cell.value = val
    cell.fill = fill(LT_RED)
    cell.font = font(DK_RED if col == "B" else DK_RED, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col == "D":
        cell.number_format = NF_DOLLAR

section_hdr(ws4, 25, "  EPS IMPACT")
label(ws4, "B26", "  After-tax financing cost ($M)")
data_cell(ws4, "D26", "=D23", nf=NF_DOLLAR)

label(ws4, "B27", "  MSFT diluted shares (M)")
ws4["D27"].value = "='2 - Standalone P&L'!C19"
ws4["D27"].alignment = align("right")
ws4["D27"].number_format = NF_DOLLAR

label(ws4, "B28", "  EPS drag from financing ($/share)")
ws4["D28"].value = "=-D23/D27"    # FIXED: removed /1000
ws4["D28"].fill = fill(LT_RED)
ws4["D28"].font = font(DK_RED, bold=True)
ws4["D28"].alignment = align("right")
ws4["D28"].number_format = NF_DOLLAR2
note(ws4, "E28", "Negative = EPS drag. Formula: −After-tax cost ÷ Diluted shares")

# Interview insight callout
insight = ws4["B30"]
insight.value = ("INTERVIEW INSIGHT: Why is interest tax-deductible? Because under GAAP and tax law, "
                 "interest paid on debt reduces taxable income. The government effectively subsidises "
                 "18.8¢ of every dollar of interest Microsoft pays. This is the 'debt tax shield' — "
                 "a core concept in capital structure theory (Modigliani-Miller).")
insight.fill = fill(LT_MINT)
insight.font = font(DK_GREEN)
insight.alignment = align("left", wrap=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Synergies
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5 - Synergies")
ws5.sheet_properties.tabColor = TAB_GREEN
ws5.column_dimensions["A"].width = 2
ws5.column_dimensions["B"].width = 38
ws5.column_dimensions["C"].width = 14
ws5.column_dimensions["D"].width = 14
ws5.column_dimensions["E"].width = 42
ws5.column_dimensions["F"].width = 12

title_row(ws5, 2, "STEP 5: SYNERGY ASSUMPTIONS & REALISATION SCHEDULE")
title_row(ws5, 3, "Annual run-rate synergies and year-by-year realisation  |  $M  |  3-year ramp",
          subtitle=True)

for coord, hdr, h in [("B5","Synergy Item","left"),("C5","Type","right"),
                       ("D5","Run-Rate ($M)","right"),("E5","Source / Rationale","right"),
                       ("F5","Confidence","right")]:
    col_header(ws5, coord, hdr, horiz=h)

section_hdr(ws5, 7, "  COST SYNERGIES — higher confidence, typically modelled at full credit")

cost_syn = [
    (8,  "  Corporate function overlap (Finance, HR, Legal, IT)", "Cost", 450, "High",
         "~3,750 roles eliminated × $120K avg total comp"),
    (9,  "  Infrastructure / cloud migration to Azure",           "Cost", 380, "High",
         "ATVI workloads move to MSFT Azure from AWS/GCP"),
    (10, "  Procurement savings (shared vendor contracts)",        "Cost", 180, "Medium",
         "Leverage MSFT's $50B+ annual procurement scale"),
    (11, "  Real estate consolidation",                           "Cost",  90, "Medium",
         "ATVI offices overlap with MSFT in LA, Seattle, Dublin"),
]
for r, lbl, typ, amt, conf, src in cost_syn:
    label(ws5, f"B{r}", lbl)
    ws5[f"C{r}"].value = typ;   ws5[f"C{r}"].alignment = align("right")
    data_cell(ws5, f"D{r}", amt, nf=NF_DOLLAR)
    note(ws5, f"E{r}", src)
    ws5[f"F{r}"].value = conf;  ws5[f"F{r}"].alignment = align("right")

# Total cost synergies (green highlight)
for col, val in [("B", "  Total Cost Synergies"), ("D", "=SUM(D8:D11)")]:
    cell = ws5[f"{col}12"]
    cell.value = val
    cell.fill = fill(LT_GREEN)
    cell.font = font(DARK_NAVY if col == "B" else BLACK, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col == "D":
        cell.number_format = NF_DOLLAR

section_hdr(ws5, 14, "  REVENUE SYNERGIES — lower confidence; apply 50% haircut in conservative case")

rev_syn = [
    (15, "  Game Pass subscriber uplift from ATVI titles",         "Revenue", 800, "Medium",
         "CoD/WoW/Overwatch on Game Pass drives net adds at $15/mo"),
    (16, "  Mobile gaming monetisation (King + MSFT distribution)","Revenue", 600, "Low",
         "Candy Crush + MSFT mobile reach + Azure backend savings"),
    (17, "  Cross-sell / upsell between player bases",             "Revenue", 350, "Low",
         "ATVI players upgrade to Xbox/MSFT ecosystem products"),
    (18, "  International market expansion (China, EM)",           "Revenue", 250, "Low",
         "ATVI titles + MSFT enterprise relationships open new markets"),
]
for r, lbl, typ, amt, conf, src in rev_syn:
    label(ws5, f"B{r}", lbl)
    ws5[f"C{r}"].value = typ;  ws5[f"C{r}"].alignment = align("right")
    data_cell(ws5, f"D{r}", amt, nf=NF_DOLLAR)
    note(ws5, f"E{r}", src)
    ws5[f"F{r}"].value = conf; ws5[f"F{r}"].alignment = align("right")

# Total revenue synergies
for col, val in [("B", "  Total Revenue Synergies (gross)"), ("D", "=SUM(D15:D18)")]:
    cell = ws5[f"{col}19"]
    cell.value = val
    cell.fill = fill(LT_GREEN)
    cell.font = font(DARK_NAVY if col == "B" else BLACK, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col == "D":
        cell.number_format = NF_DOLLAR

section_hdr(ws5, 21, "  TOTAL SYNERGIES SUMMARY")

summary = [
    (22, "  Total Cost Synergies",              "=D12"),
    (23, "  Total Revenue Synergies (gross)",   "=D19"),
]
for r, lbl, val in summary:
    label(ws5, f"B{r}", lbl)
    data_cell(ws5, f"D{r}", val, nf=NF_DOLLAR)

label(ws5, "B24", "  Revenue synergy haircut (conservative case)")
ws5["C24"].value = 0.5
ws5["C24"].alignment = align("right")
ws5["C24"].number_format = NF_PCT
note(ws5, "E24",
     "Standard practice: discount revenue synergies 50% to reflect execution risk vs. cost synergies")

label(ws5, "B25", "  Adjusted Revenue Synergies (post-haircut)")
data_cell(ws5, "D25", "=D23*(1-C24)", nf=NF_DOLLAR)

for col, val in [("B", "  TOTAL RUN-RATE SYNERGIES (base case)"), ("D", "=D22+D25")]:
    cell = ws5[f"{col}26"]
    cell.value = val
    cell.fill = fill(LT_GREEN)
    cell.font = font(DK_GREEN if col == "B" else DK_GREEN, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col == "D":
        cell.number_format = NF_DOLLAR

section_hdr(ws5, 28, "  REALISATION SCHEDULE (how quickly synergies are achieved)")

for coord, hdr in [("B29","Item"),("C29","Run-Rate"),
                   ("D29","Year 1"),("E29","Year 2"),("F29","Year 3")]:
    col_header(ws5, coord, hdr, horiz="left" if coord=="B29" else "right")

label(ws5, "B30", "  Realisation rate assumption")
for col, val in [("D30", 0.4), ("E30", 0.7), ("F30", 1.0)]:
    ws5[col].value = val
    ws5[col].alignment = align("right")
    ws5[col].number_format = NF_PCT

# Cost synergy ramp
label(ws5, "B31", "    Cost synergies achieved ($M)")
ws5["C31"].value = "=D12";  ws5["C31"].alignment = align("right");  ws5["C31"].number_format = NF_DOLLAR
for col, formula in [("D31","=C31*D30"),("E31","=C31*E30"),("F31","=C31*F30")]:
    data_cell(ws5, col, formula, nf=NF_DOLLAR)

# Revenue synergy ramp
label(ws5, "B32", "    Revenue synergies achieved ($M)")
ws5["C32"].value = "=D25";  ws5["C32"].alignment = align("right");  ws5["C32"].number_format = NF_DOLLAR
for col, formula in [("D32","=C32*D30"),("E32","=C32*E30"),("F32","=C32*F30")]:
    data_cell(ws5, col, formula, nf=NF_DOLLAR)

# Total gross synergies realised
for col, val in [("B","  Total gross synergies realised ($M)"),
                 ("C","=D26"),("D","=D31+D32"),("E","=E31+E32"),("F","=F31+F32")]:
    cell = ws5[f"{col}33"]
    cell.value = val
    cell.fill = fill(LT_GREEN)
    cell.font = font(DARK_NAVY if col == "B" else BLACK, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col != "B":
        cell.number_format = NF_DOLLAR

label(ws5, "B34", "  Tax rate (MSFT effective rate)")
ws5["C34"].value = "='4 - Financing Cost'!$C$12"   # FIXED: was C10 (interest rate)
ws5["C34"].alignment = align("right")
ws5["C34"].number_format = NF_PCT

# After-tax synergies
for col, val in [("B","  After-tax synergies ($M) — flows to pro forma NI"),
                 ("D","=D33*(1-C34)"),("E","=E33*(1-C34)"),("F","=F33*(1-C34)")]:
    cell = ws5[f"{col}35"]
    cell.value = val
    cell.fill = fill(LT_GREEN)
    cell.font = font(DK_GREEN if col == "B" else DK_GREEN, bold=True)
    cell.alignment = align("left" if col == "B" else "right")
    if col != "B":
        cell.number_format = NF_DOLLAR

# EPS uplift per share — FIXED: removed /1000
label(ws5, "B36", "  EPS uplift from synergies ($/share)")
for col, val in [("D36", "=D35/'2 - Standalone P&L'!C19"),
                 ("E36", "=E35/'2 - Standalone P&L'!C19"),
                 ("F36", "=F35/'2 - Standalone P&L'!C19")]:
    cell = ws5[col]
    cell.value = val
    cell.fill = fill(LT_GREEN)
    cell.font = font(DK_GREEN, bold=True)
    cell.alignment = align("right")
    cell.number_format = NF_DOLLAR2

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Pro Forma EPS
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6 - Pro Forma EPS")
ws6.sheet_properties.tabColor = TAB_NAVY
ws6.column_dimensions["A"].width = 2
ws6.column_dimensions["B"].width = 40
ws6.column_dimensions["C"].width = 14
ws6.column_dimensions["D"].width = 14
ws6.column_dimensions["E"].width = 14
ws6.column_dimensions["F"].width = 14
ws6.column_dimensions["G"].width = 42

title_row(ws6, 2, "STEP 6: PRO FORMA EPS & ACCRETION / DILUTION VERDICT")
title_row(ws6, 3, "Combines all prior steps into the final answer  |  $M unless noted", subtitle=True)

for coord, hdr in [("B5","Line Item"),("C5","Year 1"),("D5","Year 2"),
                   ("E5","Year 3"),("F5","Run-Rate"),("G5","Formula / Notes")]:
    col_header(ws6, coord, hdr, horiz="left" if coord=="B5" else "right")

section_hdr(ws6, 7, "  PRO FORMA NET INCOME BUILD")

# MSFT standalone NI (light blue)
for col, val, fc_c, bold_c in [
    ("B8", "  MSFT standalone net income ($M)", DARK_NAVY, True),
    ("C8", "='2 - Standalone P&L'!C17",         DARK_NAVY, True),
    ("D8", "='2 - Standalone P&L'!C17",         DARK_NAVY, True),
    ("E8", "='2 - Standalone P&L'!C17",         DARK_NAVY, True),
    ("F8", "='2 - Standalone P&L'!C17",         DARK_NAVY, True),
]:
    cell = ws6[col]
    cell.value = val
    cell.fill = fill(LT_BLUE)
    cell.font = font(fc_c, bold_c)
    cell.alignment = align("left" if col=="B8" else "right")
    if col != "B8":
        cell.number_format = NF_DOLLAR
ws6["G8"].value = "Links from Standalone P&L tab. Fixed across all years (no growth modelled here)"
ws6["G8"].fill = fill(LT_BLUE)
ws6["G8"].font = font(GRAY_BODY)
ws6["G8"].alignment = align("left", wrap=True)

# ATVI NI (light green)
for col, val in [("B9", "  (+) ATVI acquired net income ($M)"),
                 ("C9", "='2 - Standalone P&L'!D17"),
                 ("D9", "='2 - Standalone P&L'!D17"),
                 ("E9", "='2 - Standalone P&L'!D17"),
                 ("F9", "='2 - Standalone P&L'!D17")]:
    cell = ws6[col]
    cell.value = val
    cell.fill = fill(LT_GREEN)
    cell.font = font(GREEN_CELL)
    cell.alignment = align("left" if col=="B9" else "right")
    if col != "B9":
        cell.number_format = NF_DOLLAR
ws6["G9"].value = "ATVI net income added to combined entity. Assume flat for simplicity."
ws6["G9"].fill = fill(LT_GREEN)
ws6["G9"].font = font(GRAY_BODY)
ws6["G9"].alignment = align("left", wrap=True)

# After-tax financing cost (light red)
for col, val in [("B10", "  (−) After-tax financing cost ($M)"),
                 ("C10", "=-'4 - Financing Cost'!D23"),
                 ("D10", "=-'4 - Financing Cost'!D23"),
                 ("E10", "=-'4 - Financing Cost'!D23"),
                 ("F10", "=-'4 - Financing Cost'!D23")]:
    cell = ws6[col]
    cell.value = val
    cell.fill = fill(LT_RED)
    cell.font = font(RED_CELL)
    cell.alignment = align("left" if col=="B10" else "right")
    if col != "B10":
        cell.number_format = NF_DOLLAR
ws6["G10"].value = "Annual interest cost after tax shield. Fixed; reduces as debt is optionally repaid."
ws6["G10"].fill = fill(LT_RED)
ws6["G10"].font = font(GRAY_BODY)
ws6["G10"].alignment = align("left", wrap=True)

# After-tax PPA amortisation (light red)
for col, val in [("B11", "  (−) After-tax PPA amortisation ($M)"),
                 ("C11", "=-'3 - PPA & Amortisation'!C31"),
                 ("D11", "=-'3 - PPA & Amortisation'!C31"),
                 ("E11", "=-'3 - PPA & Amortisation'!C31"),
                 ("F11", "=-'3 - PPA & Amortisation'!C31")]:
    cell = ws6[col]
    cell.value = val
    cell.fill = fill(LT_RED)
    cell.font = font(RED_CELL)
    cell.alignment = align("left" if col=="B11" else "right")
    if col != "B11":
        cell.number_format = NF_DOLLAR
ws6["G11"].value = ("Declines over time as intangibles become fully amortised. "
                    "Run-Rate column shows steady-state.")
ws6["G11"].fill = fill(LT_RED)
ws6["G11"].font = font(GRAY_BODY)
ws6["G11"].alignment = align("left", wrap=True)

# After-tax synergies (light green)
for col, val in [("B12", "  (+) After-tax synergies — realised ($M)"),
                 ("C12", "='5 - Synergies'!D35"),
                 ("D12", "='5 - Synergies'!E35"),
                 ("E12", "='5 - Synergies'!F35"),
                 ("F12", "='5 - Synergies'!D26*(1-'4 - Financing Cost'!$C$12)")]:  # FIXED: C12 not C10
    cell = ws6[col]
    cell.value = val
    cell.fill = fill(LT_GREEN)
    cell.font = font(GREEN_CELL)
    cell.alignment = align("left" if col=="B12" else "right")
    if col != "B12":
        cell.number_format = NF_DOLLAR
ws6["G12"].value = "Ramps up: 40% Yr1 → 70% Yr2 → 100% Yr3 → full run-rate thereafter."
ws6["G12"].fill = fill(LT_GREEN)
ws6["G12"].font = font(GRAY_BODY)
ws6["G12"].alignment = align("left", wrap=True)

# Pro Forma NI total (light blue)
for col, val in [("B14", "  PRO FORMA NET INCOME ($M)"),
                 ("C14", "=C8+C9+C10+C11+C12"),
                 ("D14", "=D8+D9+D10+D11+D12"),
                 ("E14", "=E8+E9+E10+E11+E12"),
                 ("F14", "=F8+F9+F10+F11+F12")]:
    cell = ws6[col]
    cell.value = val
    cell.fill = fill(LT_BLUE)
    cell.font = font(DARK_NAVY, bold=True)
    cell.alignment = align("left" if col=="B14" else "right")
    if col != "B14":
        cell.number_format = NF_DOLLAR

section_hdr(ws6, 16, "  PRO FORMA EPS — THE VERDICT")

# Diluted shares
label(ws6, "B17", "  Diluted shares outstanding (M) — unchanged (all-cash)")
for col in ["C17","D17","E17","F17"]:
    ws6[col].value = "='2 - Standalone P&L'!C19"
    ws6[col].alignment = align("right")
    ws6[col].number_format = NF_DOLLAR

# Pro Forma EPS — FIXED: removed /1000
for col, val in [("B18", "  PRO FORMA EPS ($/share)"),
                 ("C18", "=C14/C17"), ("D18", "=D14/D17"),
                 ("E18", "=E14/E17"), ("F18", "=F14/F17")]:
    cell = ws6[col]
    cell.value = val
    cell.fill = fill(LT_BLUE)
    cell.font = font(DARK_NAVY, bold=True)
    cell.alignment = align("left" if col=="B18" else "right")
    if col != "B18":
        cell.number_format = NF_DOLLAR2

# Standalone EPS
label(ws6, "B19", "  MSFT standalone EPS ($/share)")
for col in ["C19","D19","E19","F19"]:
    ws6[col].value = "='2 - Standalone P&L'!C20"
    ws6[col].alignment = align("right")
    ws6[col].number_format = NF_DOLLAR2

# Accretion / Dilution $
label(ws6, "B20", "  Accretion / (Dilution) ($/share)")
for col, val in [("C20","=C18-C19"),("D20","=D18-D19"),
                 ("E20","=E18-E19"),("F20","=F18-F19")]:
    ws6[col].value = val
    ws6[col].alignment = align("right")
    ws6[col].number_format = NF_DOLLAR2

# Accretion % (amber highlight)
for col, val in [("B21", "  ACCRETION / (DILUTION) % vs. standalone"),
                 ("C21","=C20/C19"),("D21","=D20/D19"),
                 ("E21","=E20/E19"),("F21","=F20/F19")]:
    cell = ws6[col]
    cell.value = val
    cell.fill = fill(LT_AMBER)
    cell.font = font(DK_AMBER, bold=True)
    cell.alignment = align("left" if col=="B21" else "right")
    if col != "B21":
        cell.number_format = NF_PCT

# ── Sensitivity Table ──────────────────────────────────────────────────────
section_hdr(ws6, 23, "  SENSITIVITY TABLE: Year 1 EPS Accretion / (Dilution) %")

ws6["B24"].value = ("Rows = blended interest rate on new debt  |  "
                    "Columns = Year 1 gross synergies realised ($M)")
ws6["B24"].alignment = align("left", wrap=True)
ws6["B24"].font = font(GRAY_BODY)

# Header row
for coord, hdr, h in [("B25","Rate \\ Synergies →","left"),
                       ("C25","$0M","right"),("D25","$500M","right"),
                       ("E25","$1,000M","right"),("F25","$1,500M","right"),
                       ("G25","$2,000M","right"),("H25","$2,500M","right"),
                       ("I25","$3,500M","right")]:
    cell = ws6[coord]
    cell.value = hdr
    cell.fill = fill(LT_GRAY)
    cell.font = font(DARK_NAVY, bold=True)
    cell.alignment = align(h)

# Sensitivity formulas — REBUILT with correct logic
rates_sens  = [0.03, 0.04, 0.046, 0.055, 0.065]
labels_sens = ["3.0%", "4.0%", "4.6%", "5.5%", "6.5%"]
syns_sens   = [0, 500, 1000, 1500, 2000, 2500, 3500]
cols_sens   = list("CDEFGHI")

for row_i, (row_num, rate, rate_lbl) in enumerate(zip(range(26, 31), rates_sens, labels_sens)):
    cell = ws6[f"B{row_num}"]
    cell.value = rate_lbl
    cell.fill = fill(LT_GRAY)
    cell.font = font(DARK_NAVY, bold=True)
    cell.alignment = align("left")

    for col, syn in zip(cols_sens, syns_sens):
        formula = (
            f"=( ('2 - Standalone P&L'!C17+'2 - Standalone P&L'!D17"
            f"-('4 - Financing Cost'!C8*{rate}+'4 - Financing Cost'!C9*'4 - Financing Cost'!C11)"
            f"*(1-'4 - Financing Cost'!C12)"
            f"-'3 - PPA & Amortisation'!C31"
            f"+{syn}*(1-'4 - Financing Cost'!C12))"
            f"/'2 - Standalone P&L'!C19"
            f"-'2 - Standalone P&L'!C20)/'2 - Standalone P&L'!C20"
        )
        cell = ws6[f"{col}{row_num}"]
        cell.value = formula
        cell.number_format = NF_PCT
        cell.alignment = align("right")
        # Base-case cell (4.6% rate, $2,000M synergies = G28) gets navy highlight
        is_base = (rate == 0.046 and syn == 2000)
        if is_base:
            cell.fill = fill(NAVY)
            cell.font = font(WHITE, bold=True)
        else:
            cell.fill = fill(LT_GRAY)
            cell.font = font(DARK_NAVY)

ws6["B32"].value = ("Navy cell = base case. Positive % = accretive. Negative % = dilutive. "
                    "Sensitivity shows how dependent the verdict is on synergy realisation — "
                    "the most defensible assumption to challenge in an interview.")
ws6["B32"].font = font(GRAY_BODY, italic=True)
ws6["B32"].alignment = align("left", wrap=True)

# ── EPS Bridge ─────────────────────────────────────────────────────────────
section_hdr(ws6, 34, "  EPS BRIDGE (Year 1) — sources of accretion / dilution")

bridge = [
    (36, "  MSFT standalone EPS",                   LT_BLUE,  DARK_NAVY, False,
         "='2 - Standalone P&L'!C20"),
    (37, "  (+) ATVI earnings contribution",         LT_GREEN, DK_GREEN,  False,
         "='2 - Standalone P&L'!D17/'2 - Standalone P&L'!C19"),
    (38, "  (−) After-tax interest expense per share", LT_RED, DK_RED,   False,
         "=-'4 - Financing Cost'!D23/'2 - Standalone P&L'!C19"),
    (39, "  (−) After-tax PPA amortisation per share", LT_RED, DK_RED,   False,
         "=-'3 - PPA & Amortisation'!C31/'2 - Standalone P&L'!C19"),
    (40, "  (+) After-tax Yr1 synergies per share",  LT_GREEN, DK_GREEN,  False,
         "='5 - Synergies'!D35/'2 - Standalone P&L'!C19"),
    (41, "  = PRO FORMA EPS",                        LT_BLUE,  DARK_NAVY, True,
         "=C18"),
]
for r, lbl, fg_c, fc_c, bold_c, formula in bridge:
    cell_b = ws6[f"B{r}"]
    cell_b.value = lbl
    cell_b.fill = fill(fg_c)
    cell_b.font = font(fc_c, bold_c)
    cell_b.alignment = align("left")

    cell_c = ws6[f"C{r}"]
    cell_c.value = formula
    cell_c.fill = fill(fg_c)
    cell_c.font = font(fc_c, bold_c)
    cell_c.alignment = align("right")
    cell_c.number_format = NF_DOLLAR2

# ══════════════════════════════════════════════════════════════════════════════
# Global row-height defaults (all sheets)
# ══════════════════════════════════════════════════════════════════════════════
for ws in [ws0, ws1, ws2, ws3, ws4, ws5, ws6]:
    for row_num in range(1, ws.max_row + 2):
        ws.row_dimensions[row_num].height = 15.75
    ws.row_dimensions[5].height = 21.75   # column-header row

# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "MSFT_ATVI_Accretion_Dilution_Model_Generated.xlsx")
wb.save(out_path)
print(f"Saved → {out_path}")
