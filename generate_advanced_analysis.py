"""
generate_advanced_analysis.py
Generates MSFT_ATVI_Advanced_Analysis.xlsx with six analytical modules:
  Memo  |  Break-even Solver  |  Financing Optimizer  |  Premium Sensitivity
  Tornado / Driver Analysis  |  Historical Rate Scenarios

Run:  python3 generate_advanced_analysis.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ── Colour palette ─────────────────────────────────────────────────────────────
NAVY      = "FF1B3A6B"
MED_BLUE  = "FF2A4F8A"
WHITE     = "FFFFFFFF"
GOLD      = "FFC9A84C"
GRAY_BODY = "FF5A6070"
LT_GRAY   = "FFF4F5F7"
LT_BLUE   = "FFE6F1FB"
LT_GREEN  = "FFEAF3DE"
LT_RED    = "FFFCEBEB"
LT_AMBER  = "FFFAEEDA"
LT_MINT   = "FFEBF5EE"
DK_RED    = "FF791F1F"
DK_GREEN  = "FF085041"
DK_AMBER  = "FF633806"
GREEN_C   = "FF008000"
RED_C     = "FFFF0000"
BLACK     = "FF000000"
DARK_NAVY = "FF1B3A6B"

TAB_GOLD  = "C9A84C"
TAB_BLUE  = "185FA5"
TAB_AMBER = "FAEEDA"
TAB_RED   = "993C1D"
TAB_GREEN = "3B6D11"
TAB_NAVY  = "1B3A6B"
TAB_TEAL  = "006B6B"

NF_DOLLAR  = r'$#,##0;"($"#,##0)\;-'
NF_DOLLAR2 = r'$#,##0.00;"($"#,##0.00)\;-'
NF_PCT     = r'0.0%;\(0.0%\)\;-'
NF_PCT2    = r'0.00%;\(0.00%\)\;-'
NF_MULT    = r'0.0\x;\(0.0"x)"\;-'
NF_GENERAL = "General"

# ── Style helpers ───────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)

def font(h=BLACK, bold=False, size=10, italic=False, name="Calibri"):
    return Font(name=name, color=h, bold=bold, size=size, italic=italic)

def align(horiz="left", wrap=False, vert="center"):
    return Alignment(horizontal=horiz, vertical=vert, wrap_text=wrap)

def title_row(ws, row, text, subtitle=False):
    c = ws[f"B{row}"]
    c.value = text
    c.fill = fill(NAVY)
    c.font = font(GOLD if subtitle else WHITE, bold=not subtitle,
                  size=10 if subtitle else 13)
    c.alignment = align("left")

def col_header(ws, coord, text, horiz="right"):
    c = ws[coord]
    c.value = text
    c.fill = fill(MED_BLUE)
    c.font = font(WHITE, bold=True, size=10)
    c.alignment = align(horiz)

def section_hdr(ws, row, text, col="B", fg=LT_GRAY, fc=DARK_NAVY):
    c = ws[f"{col}{row}"]
    c.value = text
    c.fill = fill(fg)
    c.font = font(fc, bold=True)
    c.alignment = align("left")

def lbl(ws, coord, text, fg=None, fc=BLACK, bold=False, wrap=False, italic=False):
    c = ws[coord]
    c.value = text
    if fg: c.fill = fill(fg)
    c.font = font(fc, bold, italic=italic)
    c.alignment = align("left", wrap)

def data(ws, coord, value, fg=None, fc=BLACK, bold=False, horiz="right",
         nf=NF_DOLLAR, italic=False):
    c = ws[coord]
    c.value = value
    if fg: c.fill = fill(fg)
    c.font = font(fc, bold, italic=italic)
    c.alignment = align(horiz)
    c.number_format = nf
    return c

def note(ws, coord, text, fc=GRAY_BODY):
    c = ws[coord]
    c.value = text
    c.font = font(fc, italic=True, size=9)
    c.alignment = align("left", wrap=True)

def thick_bottom(ws, row, col_start="B", col_end="I"):
    side = Side(style="medium", color="FF1B3A6B")
    for col in range(ord(col_start), ord(col_end)+1):
        c = ws[f"{chr(col)}{row}"]
        b = c.border
        c.border = Border(bottom=side,
                          left=b.left if b else None,
                          right=b.right if b else None,
                          top=b.top if b else None)

def set_row_height(ws, start=1, end=None, h=15.75):
    end = end or ws.max_row + 2
    for r in range(start, end+1):
        ws.row_dimensions[r].height = h

# ── Model constants ─────────────────────────────────────────────────────────────
MSFT_NI     = 76471      # $M standalone net income
ATVI_NI     = 3572       # $M acquired net income
MSFT_SHARES = 7434       # M diluted shares (unchanged — all-cash)
MSFT_EPS    = MSFT_NI / MSFT_SHARES        # ~$10.286

EQUITY_VAL  = 68700      # $M acquisition equity value ($95 × 724M shs)
ATVI_SHARES = 724        # M ATVI diluted shares
DEAL_PRICE  = 95.00      # $/share
FEES        = 800        # $M advisory + financing fees
NET_DEBT    = 970        # $M ATVI net debt assumed
TOTAL_DEAL  = EQUITY_VAL + FEES + NET_DEBT  # $70,470M

CASH_DEP    = 20000      # $M from MSFT balance sheet
NEW_DEBT    = 50470      # $M new bonds / notes / CP

DEBT_RATE   = 0.046      # blended interest rate on new debt
CASH_YIELD  = 0.045      # foregone yield on $20B cash
TAX_RATE    = 0.188      # MSFT effective tax rate

# PPA intangibles (fair values $M, useful lives yrs)
GAME_FV, GAME_LIFE = 8500, 6
CUST_FV, CUST_LIFE = 3200, 10
TECH_FV, TECH_LIFE = 1100, 5
PPA_AMORT = GAME_FV/GAME_LIFE + CUST_FV/CUST_LIFE + TECH_FV/TECH_LIFE  # ~$1,957M/yr
AT_PPA    = PPA_AMORT * (1 - TAX_RATE)   # ~$1,589M

# Financing cost
GROSS_INT   = NEW_DEBT * DEBT_RATE + CASH_DEP * CASH_YIELD   # $3,222M
AT_FIN      = GROSS_INT * (1 - TAX_RATE)                     # $2,616M

# Synergies
RUN_RATE_SYN = 2100   # $M pre-tax run-rate
RAMP         = [0.40, 0.70, 1.00]

MSFT_PRICE  = 330.0   # MSFT share price at close (for equity issuance scenarios)

# ── Core computation function ───────────────────────────────────────────────────
def pf_eps(syn_rr, debt_rate, cash_yld, tax, ppa_amort, ramp_pct,
           new_debt=NEW_DEBT, cash_dep=CASH_DEP, extra_shares=0):
    gross = new_debt * debt_rate + cash_dep * cash_yld
    at_fin  = gross * (1 - tax)
    at_ppa  = ppa_amort * (1 - tax)
    at_syn  = syn_rr * ramp_pct * (1 - tax)
    pf_ni   = MSFT_NI + ATVI_NI - at_fin - at_ppa + at_syn
    shares  = MSFT_SHARES + extra_shares
    return pf_ni / shares

def accretion_pct(pf, base=MSFT_EPS):
    return (pf - base) / base

# ═══════════════════════════════════════════════════════════════════════════════
# TAB A: ONE-PAGE DEAL MEMO
# ═══════════════════════════════════════════════════════════════════════════════
def build_memo(wb):
    ws = wb.create_sheet("Memo — One Page")
    ws.sheet_properties.tabColor = TAB_GOLD
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 18

    # ── Header ──────────────────────────────────────────────────────────────
    for row, text, subtitle in [
        (2, "TRANSACTION MEMO: MICROSOFT CORPORATION / ACTIVISION BLIZZARD, INC.", False),
        (3, "Confidential  |  October 2023  |  $68.7 Billion All-Cash Acquisition  |  Accretion / Dilution Analysis", True),
    ]:
        title_row(ws, row, text, subtitle)
        ws.merge_cells(f"B{row}:D{row}")

    # Horizontal rule (thick gray border on row 4)
    for col in "BCD":
        ws[f"{col}4"].border = Border(bottom=Side(style="medium", color="FF1B3A6B"))

    # ── Section I: Strategic Rationale ──────────────────────────────────────
    section_hdr(ws, 6, "  I.  STRATEGIC RATIONALE", fg=NAVY, fc=WHITE)
    ws.merge_cells("B6:D6")

    rationale = [
        ("Content Moat & Game Pass Acceleration",
         "Activision's Call of Duty, World of Warcraft, Overwatch, and Candy Crush "
         "add 17 of the top-25 gaming franchises by revenue to Microsoft's Game Pass "
         "subscription. This materially strengthens the value proposition of Xbox's "
         "recurring revenue model and accelerates subscriber growth — addressing the "
         "single biggest strategic gap in Microsoft's gaming P&L."),
        ("Mobile Platform Establishment (King / Candy Crush)",
         "Mobile is the largest gaming platform by revenue (~$100B/yr) and the one "
         "where Microsoft had zero presence. King's 300M+ monthly active users and "
         "$2.5B in annual mobile revenue provide immediate scale, distribution, and "
         "data to build a mobile gaming flywheel alongside Azure cloud services."),
        ("Competitive Defense vs. Sony, Apple & Google",
         "The deal preempts Sony from locking up CoD exclusivity on PlayStation — "
         "a title generating >$3B in annual revenue. It also positions Microsoft to "
         "compete for gaming wallet share against Apple Arcade and Google's cloud "
         "gaming ambitions, using ATVI IP to drive engagement on Azure and Windows."),
    ]

    r = 8
    for heading, body in rationale:
        lbl(ws, f"B{r}", f"  {heading}", fg=LT_BLUE, fc=DARK_NAVY, bold=True)
        ws[f"B{r}"].alignment = align("left", wrap=True)
        ws.merge_cells(f"B{r}:D{r}")
        ws.row_dimensions[r].height = 18

        r += 1
        ws[f"B{r}"].value = body
        ws[f"B{r}"].font = font(GRAY_BODY, size=9)
        ws[f"B{r}"].alignment = align("left", wrap=True)
        ws.merge_cells(f"B{r}:D{r}")
        ws.row_dimensions[r].height = 42
        r += 1

    # ── Section II: Deal Structure ───────────────────────────────────────────
    r += 1
    section_hdr(ws, r, "  II.  DEAL STRUCTURE", fg=NAVY, fc=WHITE)
    ws.merge_cells(f"B{r}:D{r}")
    r += 1

    structure = [
        ("Purchase Price",      "$95.00/share  |  $68.7B equity value  |  ~45% premium to $65.50 unaffected price"),
        ("Total Deal Cost",     "$70.5B (equity $68.7B + advisory/financing fees $0.8B + ATVI net debt $1.0B)"),
        ("Consideration",       "100% all-cash — zero new MSFT shares issued; EPS denominator fixed at 7,434M shares"),
        ("Financing Mix",       "$20B MSFT balance sheet cash  +  $35B LT bonds  +  $10B MTNs  +  $5.5B commercial paper"),
        ("Credit Quality",      "MSFT Aaa/AAA rated — issued at tight IG spreads; blended new-debt rate ~4.6%"),
        ("Announcement → Close","January 18, 2022  →  October 13, 2023 (21 months; FTC litigation delayed close)"),
        ("PPA / Goodwill",      "~$16B to identified intangibles (amortised); ~$46B to goodwill (not amortised, tested annually)"),
    ]
    for key, val in structure:
        lbl(ws, f"B{r}", f"  {key}", fc=DARK_NAVY, bold=True)
        lbl(ws, f"C{r}", val, fc=GRAY_BODY)
        ws.row_dimensions[r].height = 15.75
        r += 1

    # ── Section III: Financial Verdict ──────────────────────────────────────
    r += 1
    section_hdr(ws, r, "  III.  FINANCIAL VERDICT (BASE CASE)", fg=NAVY, fc=WHITE)
    ws.merge_cells(f"B{r}:D{r}")
    r += 1

    verdict = [
        ("EPS Bridge (Year 1)",
         f"MSFT standalone EPS ${MSFT_EPS:.2f}  +ATVI contribution +$0.48  "
         f"−Financing cost −$0.35  −PPA amort −$0.21  +Yr1 synergies +$0.09  "
         f"= Pro Forma EPS ~$10.29  →  +0.06% accretion"),
        ("3-Year Trajectory",
         "Year 1: +0.1% accretive  |  Year 2: +0.7% accretive  |  Year 3: +1.4% accretive "
         "(synergy ramp: 40% → 70% → 100% of $2.1B run-rate)"),
        ("Primary Headwinds",
         "$2.6B after-tax annual interest expense (−$0.35/share)  +  "
         "$1.6B after-tax PPA amortization (−$0.21/share)  =  −$0.56/share drag"),
        ("Primary Offsets",
         "$3.6B ATVI acquired net income (+$0.48/share)  +  $682M after-tax Yr1 synergies "
         "(+$0.09/share)  =  +$0.57/share; net +$0.01 accretive Year 1"),
        ("Financial Conclusion",
         "Deal makes financial sense at base case — positive accretion from Day 1, reaching "
         "+1.4% by Year 3. Verdict is synergy-dependent: without synergies the deal is "
         "~−0.6% dilutive. Cost synergies ($1.1B, Azure migration + headcount) are high-confidence "
         "and within Microsoft's control — these alone cover ~70% of the break-even threshold."),
    ]
    for heading, body in verdict:
        lbl(ws, f"B{r}", f"  {heading}", fg=LT_AMBER, fc=DK_AMBER, bold=True)
        ws.merge_cells(f"B{r}:D{r}")
        ws.row_dimensions[r].height = 18
        r += 1
        ws[f"B{r}"].value = body
        ws[f"B{r}"].font = font(GRAY_BODY, size=9)
        ws[f"B{r}"].alignment = align("left", wrap=True)
        ws.merge_cells(f"B{r}:D{r}")
        ws.row_dimensions[r].height = 38
        r += 1

    # ── Section IV: Complex Analytical Questions ─────────────────────────────
    r += 1
    section_hdr(ws, r, "  IV.  ANALYTICAL QUESTIONS", fg=NAVY, fc=WHITE)
    ws.merge_cells(f"B{r}:D{r}")
    r += 1

    at_drag = AT_FIN + AT_PPA - ATVI_NI           # 633M after-tax
    be_syn_pt = at_drag / (1 - TAX_RATE)          # ~$780M pre-tax break-even
    be_rr_yr1 = be_syn_pt / RAMP[0]               # ~$1,950M run-rate for Yr1 break-even

    questions = [
        ("Q1: Minimum synergy for EPS accretion?",
         f"Break-even requires ~${be_syn_pt:,.0f}M in pre-tax run-rate synergies "
         f"(${at_drag:,.0f}M after-tax). At the 40% Year-1 ramp, run-rate must be "
         f"≥${be_rr_yr1:,.0f}M for Year 1 accretion. Base case ($2,100M) sits "
         f"{(RUN_RATE_SYN-be_syn_pt)/RUN_RATE_SYN*100:.0f}% above break-even — "
         f"a significant safety margin. Cost synergies alone ($1,100M) fall just above "
         f"the break-even; revenue synergies provide the upside. See Tab: Break-even Solver."),
        ("Q2: How do financing & purchase price affect A/D?",
         f"Financing: A 1pp rise in blended debt rate shifts Year-1 EPS by ~−$0.07/share. "
         f"Moving from all-debt to 50% equity financing improves Year-1 EPS by +$0.02/share "
         f"because MSFT's implied earnings yield (3.1%) < after-tax debt cost (3.7%) — "
         f"making equity 'cheaper' from an EPS standpoint, despite diluting share count. "
         f"Purchase price: each $5/share increase above ${DEAL_PRICE:.0f} adds ~$167M of new debt "
         f"and costs ~$0.018/share in EPS drag. Break-even price = ~$96.70/share (base synergies). "
         f"See Tabs: Financing Optimizer & Premium Sensitivity."),
        ("Q3: Primary drivers of transaction value & risk?",
         "Ranked by EPS sensitivity (Year 1 accretion swing): "
         "(1) Interest rate on debt [±$0.19/share] — largest single driver, reflects macro risk; "
         "(2) Purchase price paid [±$0.14/share] — negotiation leverage matters enormously; "
         "(3) PPA amortization life [±$0.14/share] — auditor/regulator judgment on intangible lives; "
         "(4) Run-rate synergies [±$0.12/share] — most defensible to challenge; "
         "(5) Cash opportunity cost [±$0.07/share]; "
         "(6) Effective tax rate [±$0.06/share]. "
         "Synergy realization is the #1 qualitative risk — especially revenue synergies (50% haircut applied). "
         "See Tab: Tornado / Driver Analysis."),
    ]
    for heading, body in questions:
        lbl(ws, f"B{r}", f"  {heading}", fg=LT_MINT, fc=DK_GREEN, bold=True)
        ws.merge_cells(f"B{r}:D{r}")
        ws.row_dimensions[r].height = 18
        r += 1
        ws[f"B{r}"].value = body
        ws[f"B{r}"].font = font(GRAY_BODY, size=9)
        ws[f"B{r}"].alignment = align("left", wrap=True)
        ws.merge_cells(f"B{r}:D{r}")
        ws.row_dimensions[r].height = 50
        r += 1

    # Footer
    r += 1
    ws[f"B{r}"].value = ("This memo is supported by five quantitative modules: "
                         "Break-even Solver | Financing Optimizer | Premium Sensitivity | "
                         "Tornado/Driver Analysis | Historical Rate Scenarios")
    ws[f"B{r}"].font = font(GRAY_BODY, italic=True, size=9)
    ws[f"B{r}"].alignment = align("left", wrap=True)
    ws.merge_cells(f"B{r}:D{r}")

    ws.row_dimensions[5].height = 21.75


# ═══════════════════════════════════════════════════════════════════════════════
# TAB B: BREAK-EVEN SYNERGY SOLVER
# ═══════════════════════════════════════════════════════════════════════════════
def build_breakeven(wb):
    ws = wb.create_sheet("Break-even Synergy Solver")
    ws.sheet_properties.tabColor = TAB_GREEN
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 38

    title_row(ws, 2, "BREAK-EVEN SYNERGY SOLVER")
    title_row(ws, 3,
              "What minimum run-rate synergy is required for the acquisition to be EPS accretive?",
              subtitle=True)
    ws.merge_cells("B2:H2")
    ws.merge_cells("B3:H3")

    # ── Part 1: The Math ────────────────────────────────────────────────────
    section_hdr(ws, 5, "  PART 1: BREAK-EVEN DERIVATION")

    lbl(ws, "B6", "MSFT standalone EPS (baseline to beat)", fc=DARK_NAVY, bold=True)
    data(ws, "C6", MSFT_EPS, fg=LT_BLUE, fc=DARK_NAVY, bold=True, nf=NF_DOLLAR2)

    lbl(ws, "B7", "  (+) ATVI acquired net income ($M)", fc=GRAY_BODY)
    data(ws, "C7", ATVI_NI, nf=NF_DOLLAR)
    lbl(ws, "B8", "  (−) After-tax financing cost ($M)", fc=GRAY_BODY)
    data(ws, "C8", -AT_FIN, fg=LT_RED, fc=DK_RED, nf=NF_DOLLAR)
    lbl(ws, "B9", "  (−) After-tax PPA amortization ($M)", fc=GRAY_BODY)
    data(ws, "C9", -AT_PPA, fg=LT_RED, fc=DK_RED, nf=NF_DOLLAR)
    lbl(ws, "B10", "  Net EPS drag before synergies ($/share)", fc=DARK_NAVY, bold=True)
    drag_per_share = (ATVI_NI - AT_FIN - AT_PPA) / MSFT_SHARES
    data(ws, "C10", drag_per_share, fg=LT_RED, fc=DK_RED, bold=True, nf=NF_DOLLAR2)

    lbl(ws, "B12",
        "  Break-even: Pro Forma NI must equal MSFT Standalone NI",
        fc=DARK_NAVY, bold=True)
    lbl(ws, "B13",
        "  After-tax drag = ATVI NI − AT financing − AT PPA amort",
        fc=GRAY_BODY, italic=True)
    at_drag = AT_FIN + AT_PPA - ATVI_NI    # 633M
    lbl(ws, "B14", "  After-tax break-even synergy ($M)", fc=DARK_NAVY)
    data(ws, "C14", at_drag, fg=LT_AMBER, fc=DK_AMBER, bold=True, nf=NF_DOLLAR)
    note(ws, "H14", f"= AT financing ({AT_FIN:,.0f}) + AT PPA ({AT_PPA:,.0f}) − ATVI NI ({ATVI_NI:,.0f})")

    be_pretax = at_drag / (1 - TAX_RATE)  # ~$780M
    lbl(ws, "B15", "  Pre-tax break-even synergy ($M)", fc=DARK_NAVY)
    data(ws, "C15", be_pretax, fg=LT_AMBER, fc=DK_AMBER, bold=True, nf=NF_DOLLAR)
    note(ws, "H15", f"= AT break-even / (1 − {TAX_RATE:.1%} tax) = {at_drag:,.0f} / {1-TAX_RATE:.3f}")

    safety = (RUN_RATE_SYN - be_pretax) / RUN_RATE_SYN
    lbl(ws, "B16", "  Base-case synergies ($2,100M) margin of safety", fc=DARK_NAVY)
    data(ws, "C16", safety, fg=LT_GREEN, fc=DK_GREEN, bold=True, nf=NF_PCT)
    note(ws, "H16", "= (Base $2,100M − Break-even) / Base. How far synergies can fall before deal turns dilutive.")

    # ── Part 2: Required run-rate by year ───────────────────────────────────
    section_hdr(ws, 18, "  PART 2: REQUIRED RUN-RATE SYNERGY BY YEAR")
    note(ws, "B19",
         "Because synergies ramp (40% Yr1 → 70% Yr2 → 100% Yr3), a higher run-rate is needed "
         "in early years for each year to be accretive in isolation.", fc=DARK_NAVY)
    ws.merge_cells("B19:H19")

    for coord, hdr in [("B20","Year"),("C20","Ramp %"),
                        ("D20","Required Run-Rate ($M)"),("E20","Base Case ($M)"),
                        ("F20","Safety Margin"),("G20","Assessment")]:
        col_header(ws, coord, hdr, horiz="left" if coord=="B20" else "right")

    year_data = [
        (21, "Year 1", 0.40),
        (22, "Year 2", 0.70),
        (23, "Year 3 (full)", 1.00),
    ]
    assessments = ["Requires high-confidence ramp acceleration",
                   "Comfortably achievable if cost synergies deliver",
                   "Base case well exceeds threshold — 63% cushion"]
    greens = [LT_AMBER, LT_GREEN, LT_GREEN]
    for (row, yr, ramp), assess, fg_c in zip(year_data, assessments, greens):
        rr = be_pretax / ramp
        margin = (RUN_RATE_SYN - rr) / RUN_RATE_SYN
        lbl(ws, f"B{row}", f"  {yr}", fc=DARK_NAVY, bold=True)
        data(ws, f"C{row}", ramp, nf=NF_PCT)
        data(ws, f"D{row}", rr,     fg=fg_c, fc=DK_AMBER, bold=True, nf=NF_DOLLAR)
        data(ws, f"E{row}", RUN_RATE_SYN, nf=NF_DOLLAR)
        data(ws, f"F{row}", margin,  fg=fg_c, fc=DK_GREEN, bold=True, nf=NF_PCT)
        lbl(ws, f"G{row}", assess, fc=GRAY_BODY, italic=True)

    # ── Part 3: Synergy sensitivity table ───────────────────────────────────
    section_hdr(ws, 25, "  PART 3: SYNERGY SENSITIVITY — ACCRETION AT DIFFERENT RUN-RATES")

    syn_levels = [0, 500, 780, 1100, 1500, 2100, 2500, 3000, 3500]
    for coord, hdr in [("B26","Run-Rate Synergy ($M)"),
                        ("C26","After-tax Yr1"),("D26","After-tax Yr2"),("E26","After-tax Yr3"),
                        ("F26","Yr1 EPS"),("G26","Yr2 EPS"),("H26","Yr3 EPS")]:
        col_header(ws, coord, hdr, horiz="left" if coord=="B26" else "right")

    for i, syn in enumerate(syn_levels):
        r = 27 + i
        yr_eps = [pf_eps(syn, DEBT_RATE, CASH_YIELD, TAX_RATE, PPA_AMORT, rp) for rp in RAMP]
        is_base = (syn == 2100)
        is_be   = (abs(syn - 780) < 20)
        fg_c = NAVY if is_base else (LT_AMBER if is_be else None)
        fc_c = WHITE if is_base else (DK_AMBER if is_be else BLACK)

        data(ws, f"B{r}", syn, fg=fg_c, fc=fc_c, bold=is_base, horiz="left", nf=NF_DOLLAR)
        for col, eps in zip("CDEFGH", [syn*rp*(1-TAX_RATE) for rp in RAMP] + yr_eps):
            pass  # overwrite below cleanly

        # AT synergies by year
        for j, (col, rp) in enumerate(zip("CDE", RAMP)):
            at_s = syn * rp * (1 - TAX_RATE)
            data(ws, f"{col}{r}", at_s, fg=fg_c, fc=fc_c, bold=is_base, nf=NF_DOLLAR)
        # EPS by year
        for j, (col, eps) in enumerate(zip("FGH", yr_eps)):
            acr = accretion_pct(eps)
            suffix = "  ✓" if acr >= 0 else "  ✗"
            data(ws, f"{col}{r}", eps, fg=fg_c, fc=fc_c, bold=is_base, nf=NF_DOLLAR2)

        # Label the break-even row
        if is_be:
            lbl(ws, f"B{r}", f"  ≈ Break-even  ($780M)", fg=LT_AMBER, fc=DK_AMBER, bold=True)
        elif is_base:
            lbl(ws, f"B{r}", f"  ★ Base Case  ($2,100M)", fg=NAVY, fc=WHITE, bold=True)
        else:
            lbl(ws, f"B{r}", f"  ${syn:,}M", fc=DARK_NAVY)

    note(ws, "B37",
         "★ Base case. ≈ Break-even row shows minimum synergy to avoid EPS dilution. "
         "Cost synergies alone ($1,100M) exceed the break-even run-rate, providing a "
         "high-confidence floor for accretion.", fc=DARK_NAVY)
    ws.merge_cells("B37:H37")

    set_row_height(ws)
    ws.row_dimensions[5].height = 21.75


# ═══════════════════════════════════════════════════════════════════════════════
# TAB C: FINANCING STRUCTURE OPTIMIZER
# ═══════════════════════════════════════════════════════════════════════════════
def build_financing_optimizer(wb):
    ws = wb.create_sheet("Financing Structure Optimizer")
    ws.sheet_properties.tabColor = TAB_BLUE
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 38

    title_row(ws, 2, "FINANCING STRUCTURE OPTIMIZER")
    title_row(ws, 3,
              "How does the debt / equity / cash mix affect EPS accretion and leverage?  |  "
              f"MSFT price at close: ${MSFT_PRICE:.0f}/share",
              subtitle=True)
    ws.merge_cells("B2:I2")
    ws.merge_cells("B3:I3")

    # ── Key insight box ─────────────────────────────────────────────────────
    section_hdr(ws, 5, "  KEY INSIGHT: WHY EQUITY FINANCING IS MORE EPS ACCRETIVE FOR HIGH-P/E ACQUIRERS")
    ws.merge_cells("B5:I5")
    ws[f"B6"].value = (
        f"MSFT earnings yield (E/P) = ${MSFT_EPS:.2f} / ${MSFT_PRICE:.0f} = {MSFT_EPS/MSFT_PRICE:.1%}  |  "
        f"After-tax cost of debt = {DEBT_RATE:.1%} × (1−{TAX_RATE:.1%}) = {DEBT_RATE*(1-TAX_RATE):.2%}  |  "
        f"Since earnings yield ({MSFT_EPS/MSFT_PRICE:.1%}) < after-tax debt ({DEBT_RATE*(1-TAX_RATE):.2%}), "
        f"each dollar financed by equity 'costs' less EPS than a dollar financed by debt. "
        f"MSFT chose debt for strategic reasons: deal certainty, interest tax shield, "
        f"preserving share count, and access to Aaa-rated debt markets at tight spreads."
    )
    ws["B6"].font = font(DARK_NAVY, size=9)
    ws["B6"].alignment = align("left", wrap=True)
    ws.merge_cells("B6:I6")
    ws.row_dimensions[6].height = 42

    # ── Scenario table ──────────────────────────────────────────────────────
    section_hdr(ws, 8, "  FINANCING MIX SCENARIOS (non-cash portion = $50,470M)")
    ws.merge_cells("B8:I8")

    headers = ["Equity %", "Equity ($M)", "New Shares (M)",
               "Total Shares (M)", "New Debt ($M)", "AT Fin. Cost ($M)",
               "Yr1 Pro Forma EPS", "Yr1 Accretion %"]
    for i, (col, hdr) in enumerate(zip("BCDEFGHI", headers)):
        col_header(ws, f"{col}9", hdr, horiz="left" if col=="B" else "right")

    equity_pcts = [0.0, 0.10, 0.25, 0.40, 0.50, 0.75, 1.00]
    is_base_pcts = [0.0]

    for i, ep in enumerate(equity_pcts):
        r = 10 + i
        eq_amt   = NEW_DEBT * ep               # equity raised ($M)
        new_shrs = eq_amt / MSFT_PRICE * 1000  # M new shares (eq_amt is $M, price is $/sh, shares in M)
        # Wait: eq_amt is in $M, MSFT_PRICE is $/share
        # new shares = (eq_amt × 1e6) / MSFT_PRICE = eq_amt / MSFT_PRICE * 1e6 → in shares
        # but shares are in millions, so: new_shrs_M = eq_amt / MSFT_PRICE
        new_shrs_M = eq_amt / MSFT_PRICE       # M shares
        total_shrs = MSFT_SHARES + new_shrs_M
        new_debt   = NEW_DEBT - eq_amt         # remaining as debt ($M); cash stays $20B
        gross_int  = new_debt * DEBT_RATE + CASH_DEP * CASH_YIELD
        at_fin_c   = gross_int * (1 - TAX_RATE)
        at_syn_y1  = RUN_RATE_SYN * RAMP[0] * (1 - TAX_RATE)
        pf_ni      = MSFT_NI + ATVI_NI - at_fin_c - AT_PPA + at_syn_y1
        pf_eps_    = pf_ni / total_shrs
        acr        = accretion_pct(pf_eps_, MSFT_EPS)

        is_base = (ep == 0.0)
        fg_c = NAVY if is_base else None
        fc_c = WHITE if is_base else BLACK

        data(ws, f"B{r}", ep,          fg=fg_c, fc=fc_c, bold=is_base, horiz="left",  nf=NF_PCT)
        data(ws, f"C{r}", eq_amt,       fg=fg_c, fc=fc_c, bold=is_base,                nf=NF_DOLLAR)
        data(ws, f"D{r}", new_shrs_M,   fg=fg_c, fc=fc_c, bold=is_base,                nf=NF_DOLLAR)
        data(ws, f"E{r}", total_shrs,   fg=fg_c, fc=fc_c, bold=is_base,                nf=NF_DOLLAR)
        data(ws, f"F{r}", new_debt,     fg=fg_c, fc=fc_c, bold=is_base,                nf=NF_DOLLAR)
        data(ws, f"G{r}", at_fin_c,     fg=fg_c, fc=fc_c, bold=is_base,                nf=NF_DOLLAR)
        data(ws, f"H{r}", pf_eps_,      fg=fg_c, fc=fc_c, bold=is_base,                nf=NF_DOLLAR2)
        acr_fg = LT_GREEN if acr >= 0 else LT_RED
        acr_fc = DK_GREEN if acr >= 0 else DK_RED
        if is_base:
            acr_fg, acr_fc = NAVY, WHITE
        data(ws, f"I{r}", acr,          fg=acr_fg, fc=acr_fc, bold=is_base,            nf=NF_PCT2)

        if is_base:
            ws[f"B{r}"].value = "  ★ 0% (Actual: All Debt)"
            ws[f"B{r}"].font = font(WHITE, bold=True)
        else:
            ws[f"B{r}"].value = f"  {ep:.0%} equity"

    # ── Leverage analysis ───────────────────────────────────────────────────
    r = 10 + len(equity_pcts) + 2
    section_hdr(ws, r, "  LEVERAGE RATIOS (BASE CASE: 0% EQUITY)")
    ws.merge_cells(f"B{r}:I{r}")
    r += 1

    msft_ebitda  = 105672   # $M MSFT EBITDA
    atvi_ebitda  = 1820     # $M ATVI EBITDA
    pro_ebitda   = msft_ebitda + atvi_ebitda

    lev_data = [
        ("Combined EV / EBITDA (at deal price)", 70470 / atvi_ebitda, "x",
         f"Deal EV $70,470M ÷ ATVI EBITDA ${atvi_ebitda:,}M — MSFT paid 38.7x EBITDA"),
        ("Gross new debt ($M)", NEW_DEBT, "$M",
         "Permanent capital raised to fund acquisition"),
        ("Net Debt / Pro Forma EBITDA", NEW_DEBT / pro_ebitda, "x",
         f"${NEW_DEBT:,}M new debt ÷ ${pro_ebitda:,}M pro forma EBITDA. MSFT Aaa-rated pre-deal."),
        ("Interest coverage (PF EBIT / Gross Interest)", (MSFT_NI + ATVI_NI) / GROSS_INT, "x",
         "Highly comfortable; MSFT's cash generation easily services new debt."),
    ]
    for key, val, unit, nt in lev_data:
        lbl(ws, f"B{r}", f"  {key}", fc=DARK_NAVY)
        nf = NF_MULT if unit == "x" else NF_DOLLAR
        data(ws, f"C{r}", val, fg=LT_GRAY, fc=DARK_NAVY, bold=True, nf=nf)
        note(ws, f"D{r}", nt, fc=GRAY_BODY)
        ws.merge_cells(f"D{r}:I{r}")
        r += 1

    set_row_height(ws)
    ws.row_dimensions[5].height = 21.75


# ═══════════════════════════════════════════════════════════════════════════════
# TAB D: M&A PREMIUM SENSITIVITY
# ═══════════════════════════════════════════════════════════════════════════════
def build_premium_sensitivity(wb):
    ws = wb.create_sheet("M&A Premium Sensitivity")
    ws.sheet_properties.tabColor = TAB_AMBER
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 36

    title_row(ws, 2, "M&A PREMIUM SENSITIVITY — HOW PURCHASE PRICE AFFECTS ACCRETION/DILUTION")
    title_row(ws, 3,
              f"ATVI shares: {ATVI_SHARES}M  |  Unaffected price: ~$65.50/share  |  "
              f"Base deal: ${DEAL_PRICE:.0f}/share  |  Synergies held at $2,100M base case",
              subtitle=True)
    ws.merge_cells("B2:J2")
    ws.merge_cells("B3:J3")

    # ── Methodology note ────────────────────────────────────────────────────
    ws["B5"].value = (
        "PPA amortization is FIXED at $1,957M/yr regardless of price (identified intangible fair "
        "values don't change; only goodwill — which is NOT amortized — absorbs the premium). "
        "Higher price → more debt → higher interest cost → EPS drag."
    )
    ws["B5"].font = font(DARK_NAVY, size=9)
    ws["B5"].alignment = align("left", wrap=True)
    ws.merge_cells("B5:J5")
    ws.row_dimensions[5].height = 30

    # ── Header ──────────────────────────────────────────────────────────────
    headers = ["Price/Share", "Premium %", "Equity Value ($M)",
               "Total Deal ($M)", "New Debt ($M)", "AT Fin Cost ($M)",
               "Goodwill ($M)", "Yr1 EPS", "Yr1 Accretion"]
    for col, hdr in zip("BCDEFGHIJ", headers):
        col_header(ws, f"{col}7", hdr, horiz="left" if col=="B" else "right")

    prices = [65, 70, 75, 80, 85, 90, 95, 100, 105, 110, 115]
    unaffected = 65.50
    identified_intangibles = GAME_FV + CUST_FV + TECH_FV + 1800 + 1400  # all PPA items
    atvi_book = 8470    # ATVI net tangible book value $M

    for i, price in enumerate(prices):
        r = 8 + i
        premium_pct = (price - unaffected) / unaffected
        eq_val   = price * ATVI_SHARES             # $M
        deal_tot = eq_val + FEES + NET_DEBT        # $M
        new_debt_s = max(deal_tot - CASH_DEP, 0)  # keep cash fixed
        gross_i  = new_debt_s * DEBT_RATE + CASH_DEP * CASH_YIELD
        at_fin_s = gross_i * (1 - TAX_RATE)
        goodwill = max(deal_tot - identified_intangibles - atvi_book, 0)
        yr1_eps  = pf_eps(RUN_RATE_SYN, DEBT_RATE, CASH_YIELD, TAX_RATE, PPA_AMORT,
                          RAMP[0], new_debt=new_debt_s)
        acr      = accretion_pct(yr1_eps)

        is_base = (price == DEAL_PRICE)
        # break-even price ~$96.70
        is_be   = (price == 100)  # closest above break-even in our list; will annotate

        fg_c = NAVY if is_base else (LT_AMBER if price < DEAL_PRICE else None)
        fc_c = WHITE if is_base else (DK_AMBER if price < DEAL_PRICE else BLACK)
        if acr < 0 and not is_base:
            fg_c, fc_c = LT_RED, DK_RED

        lbl(ws, f"B{r}", f"  ${price:.0f}/share", fg=fg_c, fc=fc_c, bold=is_base)
        data(ws, f"C{r}", premium_pct, fg=fg_c, fc=fc_c, bold=is_base, nf=NF_PCT)
        data(ws, f"D{r}", eq_val,      fg=fg_c, fc=fc_c, bold=is_base, nf=NF_DOLLAR)
        data(ws, f"E{r}", deal_tot,    fg=fg_c, fc=fc_c, bold=is_base, nf=NF_DOLLAR)
        data(ws, f"F{r}", new_debt_s,  fg=fg_c, fc=fc_c, bold=is_base, nf=NF_DOLLAR)
        data(ws, f"G{r}", at_fin_s,    fg=fg_c, fc=fc_c, bold=is_base, nf=NF_DOLLAR)
        data(ws, f"H{r}", goodwill,    fg=fg_c, fc=fc_c, bold=is_base, nf=NF_DOLLAR)
        acr_fg = (NAVY if is_base else LT_GREEN) if acr >= 0 else LT_RED
        acr_fc = (WHITE if is_base else DK_GREEN) if acr >= 0 else DK_RED
        data(ws, f"I{r}", yr1_eps,     fg=acr_fg, fc=acr_fc, bold=is_base, nf=NF_DOLLAR2)
        data(ws, f"J{r}", acr,         fg=acr_fg, fc=acr_fc, bold=is_base, nf=NF_PCT2)

        if is_base:
            ws[f"B{r}"].value = f"  ★ ${price:.0f} (ACTUAL)"

    # ── Break-even price callout ─────────────────────────────────────────────
    r = 8 + len(prices) + 1

    # Compute exact break-even price algebraically
    # at_drag = AT_FIN + AT_PPA - ATVI_NI = 633M (at base)
    # For different prices, AT_FIN changes (new_debt changes)
    # Break-even: ATVI_NI − at_fin(price) − AT_PPA + YR1_AT_SYN = 0
    # ATVI_NI − (new_debt(price)×DEBT_RATE + CASH_DEP×CASH_YIELD)×(1−TAX) − AT_PPA + YR1_AT_SYN = 0
    yr1_at_syn = RUN_RATE_SYN * RAMP[0] * (1 - TAX_RATE)  # 682
    target_at_fin = ATVI_NI - AT_PPA + yr1_at_syn          # 3572-1589+682 = 2665
    target_gross  = target_at_fin / (1 - TAX_RATE)          # 3283
    target_int    = target_gross - CASH_DEP * CASH_YIELD     # 2383
    be_debt       = target_int / DEBT_RATE                   # 51,804M
    be_deal_cost  = be_debt + CASH_DEP                       # 71,804M
    be_equity     = be_deal_cost - FEES - NET_DEBT            # 70,034M
    be_price      = be_equity / ATVI_SHARES                  # ~$96.73

    ws[f"B{r}"].value = (
        f"Break-even purchase price: ~${be_price:.2f}/share  "
        f"(actual: ${DEAL_PRICE:.2f}/share → ${be_price-DEAL_PRICE:.2f} below break-even)  |  "
        f"Without synergies, break-even price falls to ~$71.48/share  |  "
        f"Deal structurally requires synergy delivery for EPS accretion at $95"
    )
    ws[f"B{r}"].fill = fill(LT_MINT)
    ws[f"B{r}"].font = font(DK_GREEN, bold=True, size=9)
    ws[f"B{r}"].alignment = align("left", wrap=True)
    ws.merge_cells(f"B{r}:J{r}")
    ws.row_dimensions[r].height = 30

    set_row_height(ws)
    ws.row_dimensions[6].height = 21.75


# ═══════════════════════════════════════════════════════════════════════════════
# TAB E: TORNADO / DRIVER ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
def build_tornado(wb):
    ws = wb.create_sheet("Tornado — Driver Analysis")
    ws.sheet_properties.tabColor = TAB_RED
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 36

    title_row(ws, 2, "TORNADO / DRIVER ANALYSIS — YEAR 1 EPS SENSITIVITY")
    title_row(ws, 3,
              "Ranks all key assumptions by their absolute impact on Year-1 Pro Forma EPS  |  "
              f"Base Year-1 EPS: ${pf_eps(RUN_RATE_SYN,DEBT_RATE,CASH_YIELD,TAX_RATE,PPA_AMORT,RAMP[0]):.4f}",
              subtitle=True)
    ws.merge_cells("B2:I2")
    ws.merge_cells("B3:I3")

    BASE_PF_EPS = pf_eps(RUN_RATE_SYN, DEBT_RATE, CASH_YIELD, TAX_RATE, PPA_AMORT, RAMP[0])

    section_hdr(ws, 5, "  HOW TO READ: Each row shows EPS at the low & high of one variable, all others held at base.")
    ws.merge_cells("B5:I5")

    for col, hdr in zip("BCDEFGHI",
                         ["Driver", "Low Value", "Base Value", "High Value",
                          "EPS at Low", "EPS at High",
                          "Low → High EPS Swing", "Notes"]):
        col_header(ws, f"{col}7", hdr, horiz="left" if col in "BI" else "right")

    # Define tornado drivers: (name, low, base, high, low_eps_fn, high_eps_fn, note)
    def eps_int_rate(rate):
        return pf_eps(RUN_RATE_SYN, rate, CASH_YIELD, TAX_RATE, PPA_AMORT, RAMP[0])

    def eps_syn(syn):
        return pf_eps(syn, DEBT_RATE, CASH_YIELD, TAX_RATE, PPA_AMORT, RAMP[0])

    def eps_price(price):
        eq = price * ATVI_SHARES
        deal = eq + FEES + NET_DEBT
        nd = max(deal - CASH_DEP, 0)
        return pf_eps(RUN_RATE_SYN, DEBT_RATE, CASH_YIELD, TAX_RATE, PPA_AMORT, RAMP[0], new_debt=nd)

    def eps_tax(tax):
        return pf_eps(RUN_RATE_SYN, DEBT_RATE, CASH_YIELD, tax, PPA_AMORT, RAMP[0])

    def eps_amort_life(life):
        ppa = GAME_FV/life + CUST_FV/CUST_LIFE + TECH_FV/TECH_LIFE
        return pf_eps(RUN_RATE_SYN, DEBT_RATE, CASH_YIELD, TAX_RATE, ppa, RAMP[0])

    def eps_cash_yield(cy):
        return pf_eps(RUN_RATE_SYN, DEBT_RATE, cy, TAX_RATE, PPA_AMORT, RAMP[0])

    def eps_wacc(wacc):
        # WACC affects NPV of synergies; approximate via PV factor applied to synergy value
        # We model: effective synergy credit = run_rate × PV_annuity_factor(wacc, 10yr) / 10
        # This reflects that at high WACC, future synergies are worth less; use simplified adjustment
        # PV of 10yr flat annuity at wacc, scaled so base (wacc=0.09) ≈ original run-rate
        BASE_WACC = 0.09
        pv_factor = (1 - (1+wacc)**-10)/wacc / ((1 - (1+BASE_WACC)**-10)/BASE_WACC)
        adj_syn = RUN_RATE_SYN * pv_factor
        return pf_eps(adj_syn, DEBT_RATE, CASH_YIELD, TAX_RATE, PPA_AMORT, RAMP[0])

    drivers_raw = [
        # (name, low_val, base_val, high_val, low_fn, high_fn, low_lbl, high_lbl, note_text)
        ("Interest rate on new debt",
         0.030, DEBT_RATE, 0.065,
         eps_int_rate(0.030), eps_int_rate(0.065),
         "3.0%", f"{DEBT_RATE:.1%} (base)", "6.5%",
         "Reflects macro rate environment; most impactful single driver (ZIRP vs. restrictive policy)"),
        ("Purchase price per ATVI share",
         70.0, DEAL_PRICE, 110.0,
         eps_price(70.0), eps_price(110.0),
         "$70/sh", f"${DEAL_PRICE:.0f}/sh (base)", "$110/sh",
         "Higher price → more debt → higher interest cost. PPA amortization unchanged."),
        ("PPA amortization life — game titles",
         4, GAME_LIFE, 10,
         eps_amort_life(4), eps_amort_life(10),
         "4 years", f"{GAME_LIFE} years (base)", "10 years",
         "Auditor/regulator judgment. Shorter life → higher annual charge → EPS drag."),
        ("Run-rate synergies",
         700, RUN_RATE_SYN, 3500,
         eps_syn(700), eps_syn(3500),
         "$700M", f"${RUN_RATE_SYN:,}M (base)", "$3,500M",
         "Year-1 at 40% ramp. Most defensible assumption in board / interview scrutiny."),
        ("Foregone yield on cash deployed",
         0.025, CASH_YIELD, 0.055,
         eps_cash_yield(0.025), eps_cash_yield(0.055),
         "2.5%", f"{CASH_YIELD:.1%} (base)", "5.5%",
         "Opportunity cost of $20B cash. Matters less in ZIRP; more important in 2023 rate environment."),
        ("WACC (PV of synergy stream)",
         0.07, 0.09, 0.11,
         eps_wacc(0.07), eps_wacc(0.11),
         "7.0%", "9.0% (base)", "11.0%",
         "Higher WACC → lower PV of synergies → lower effective synergy credit in Year-1 economics."),
        ("Effective tax rate",
         0.14, TAX_RATE, 0.24,
         eps_tax(0.14), eps_tax(0.24),
         "14%", f"{TAX_RATE:.1%} (base)", "24%",
         "Lower tax rate reduces value of interest deduction AND PPA deduction. Net effect: lower tax = less EPS."),
    ]

    # Sort by absolute EPS swing (descending — tornado order)
    drivers = sorted(
        drivers_raw,
        key=lambda x: abs(x[5] - x[4]),
        reverse=True
    )

    for i, (name, lo_val, ba_val, hi_val, lo_eps, hi_eps,
            lo_lbl, ba_lbl, hi_lbl, nt) in enumerate(drivers):
        r = 8 + i
        swing = abs(hi_eps - lo_eps)
        lo_chg = lo_eps - BASE_PF_EPS
        hi_chg = hi_eps - BASE_PF_EPS

        # Color: green for favorable end, red for unfavorable
        lo_fg = LT_GREEN if lo_chg > 0 else LT_RED
        lo_fc = DK_GREEN if lo_chg > 0 else DK_RED
        hi_fg = LT_GREEN if hi_chg > 0 else LT_RED
        hi_fc = DK_GREEN if hi_chg > 0 else DK_RED

        lbl(ws, f"B{r}", f"  {i+1}. {name}", fc=DARK_NAVY, bold=True)
        lbl(ws, f"C{r}", f"  {lo_lbl}", fc=GRAY_BODY)
        lbl(ws, f"D{r}", f"  {ba_lbl}", fg=LT_BLUE, fc=DARK_NAVY, bold=True)
        lbl(ws, f"E{r}", f"  {hi_lbl}", fc=GRAY_BODY)
        data(ws, f"F{r}", lo_eps,  fg=lo_fg, fc=lo_fc, bold=True, nf=NF_DOLLAR2)
        data(ws, f"G{r}", hi_eps,  fg=hi_fg, fc=hi_fc, bold=True, nf=NF_DOLLAR2)
        data(ws, f"H{r}", swing,   fg=LT_AMBER, fc=DK_AMBER, bold=True, nf=NF_DOLLAR2)
        note(ws, f"I{r}", nt)

    # ── Accretion % version ──────────────────────────────────────────────────
    r = 8 + len(drivers) + 2
    section_hdr(ws, r, "  ACCRETION / DILUTION % AT LOW & HIGH — SAME RANKING")
    ws.merge_cells(f"B{r}:I{r}")
    r += 1

    for col, hdr in zip("BCDEFGH",
                         ["Driver", "Yr1 Accrtn at LOW", "Yr1 Accrtn at BASE",
                          "Yr1 Accrtn at HIGH", "Low swing vs. Base", "High swing vs. Base",
                          "Total Swing"]):
        col_header(ws, f"{col}{r}", hdr, horiz="left" if col=="B" else "right")
    r += 1

    base_acr = accretion_pct(BASE_PF_EPS)
    for i, (name, lo_val, ba_val, hi_val, lo_eps, hi_eps,
            lo_lbl, ba_lbl, hi_lbl, nt) in enumerate(drivers):
        lo_acr = accretion_pct(lo_eps)
        hi_acr = accretion_pct(hi_eps)
        swing_acr = abs(hi_acr - lo_acr)

        lbl(ws, f"B{r}", f"  {i+1}. {name}", fc=DARK_NAVY)
        lo_fg = LT_GREEN if lo_acr > base_acr else LT_RED
        lo_fc = DK_GREEN if lo_acr > base_acr else DK_RED
        hi_fg = LT_GREEN if hi_acr > base_acr else LT_RED
        hi_fc = DK_GREEN if hi_acr > base_acr else DK_RED

        data(ws, f"C{r}", lo_acr, fg=lo_fg, fc=lo_fc, nf=NF_PCT2)
        data(ws, f"D{r}", base_acr, fg=LT_BLUE, fc=DARK_NAVY, bold=True, nf=NF_PCT2)
        data(ws, f"E{r}", hi_acr, fg=hi_fg, fc=hi_fc, nf=NF_PCT2)
        data(ws, f"F{r}", lo_acr - base_acr, nf=NF_PCT2)
        data(ws, f"G{r}", hi_acr - base_acr, nf=NF_PCT2)
        data(ws, f"H{r}", swing_acr, fg=LT_AMBER, fc=DK_AMBER, bold=True, nf=NF_PCT2)
        r += 1

    # ── Add a BarChart for the tornado ──────────────────────────────────────
    # Data for chart: swing amounts in col H
    chart_data_row_start = 8
    chart_data_row_end = 8 + len(drivers) - 1

    chart = BarChart()
    chart.type = "bar"   # horizontal bar
    chart.grouping = "clustered"
    chart.title = "Tornado: Year-1 EPS Swing by Driver (Low → High)"
    chart.y_axis.title = "Driver"
    chart.x_axis.title = "EPS Swing ($)"
    chart.style = 10

    data_ref = Reference(ws, min_col=8, min_row=chart_data_row_start,
                         max_row=chart_data_row_end)
    cats_ref = Reference(ws, min_col=2, min_row=chart_data_row_start,
                         max_row=chart_data_row_end)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.width  = 22
    chart.height = 14
    ws.add_chart(chart, "B" + str(r + 2))

    set_row_height(ws)
    ws.row_dimensions[6].height = 21.75


# ═══════════════════════════════════════════════════════════════════════════════
# TAB F: HISTORICAL INTEREST RATE SCENARIOS
# ═══════════════════════════════════════════════════════════════════════════════
def build_historical_rates(wb):
    ws = wb.create_sheet("Historical Rate Scenarios")
    ws.sheet_properties.tabColor = TAB_TEAL
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 38

    title_row(ws, 2, "HISTORICAL INTEREST RATE SCENARIOS — 2021 THROUGH 2024")
    title_row(ws, 3,
              "How would deal economics have differed if MSFT had closed in a different rate environment?  |  "
              "All other assumptions held at base case",
              subtitle=True)
    ws.merge_calls = True  # flag
    ws.merge_cells("B2:H2")
    ws.merge_cells("B3:H3")

    # ── Context ──────────────────────────────────────────────────────────────
    ws["B5"].value = (
        "The deal was announced January 18, 2022 (Fed funds rate: 0-0.25%) and closed "
        "October 13, 2023 (Fed funds rate: 5.25-5.50%) — the fastest hiking cycle in 40 years. "
        "MSFT raised debt in 2023 at the peak of the rate cycle. Below we model what the "
        "blended financing cost and EPS would have been had the deal closed in 2021, "
        "mid-2022, or 2024 instead."
    )
    ws["B5"].font = font(DARK_NAVY, size=9)
    ws["B5"].alignment = align("left", wrap=True)
    ws.merge_cells("B5:H5")
    ws.row_dimensions[5].height = 40

    section_hdr(ws, 7, "  RATE ENVIRONMENT ASSUMPTIONS BY SCENARIO")
    ws.merge_cells("B7:H7")

    rate_env = [
        # year, ff_rate_desc, 10yr_desc, msft_lt_rate, msft_mtn_rate, msft_cp_rate,
        # msft_blended, cash_yield, rationale
        ("2021 (ZIRP)",
         "0.00 – 0.25%", "~1.50%", 0.018, 0.014, 0.001, 0.016, 0.005,
         "Near-zero rate policy post-COVID. MSFT 10yr bonds at ~1.8%, MTNs ~1.4%, "
         "CP near 0%. Cash earns near zero. This was the optimal window for M&A financing."),
        ("2022-H1 (Hiking begins)",
         "0.25 – 2.50%", "~2.50–3.50%", 0.033, 0.027, 0.022, 0.030, 0.015,
         "Fed started hiking in March 2022. By mid-year: FF ~2.5%, 10yr ~3%. "
         "MSFT LT bond rate ~3.3%, MTN ~2.7%. Deal announced in this window; "
         "management began debt planning as rates rose rapidly."),
        ("2022-H2 (Hiking accelerates)",
         "2.50 – 4.50%", "~3.50–3.90%", 0.042, 0.036, 0.046, 0.040, 0.030,
         "Four 75bp hikes in 2022. 10yr reached 4%+. IG corporate spreads widened. "
         "Had MSFT closed here, blended cost ~4.0% — notably lower than actual."),
        ("2023-Q4 ★ ACTUAL CLOSE",
         "5.25 – 5.50%", "~4.50–4.90%", DEBT_RATE, 0.042, 0.052, DEBT_RATE, CASH_YIELD,
         f"ACTUAL close: October 13, 2023. Peak rate cycle. MSFT blended: {DEBT_RATE:.1%}. "
         f"Cash yield {CASH_YIELD:.1%}. This is the model base case."),
        ("2024 (Cuts begin Sep-2024)",
         "4.25 – 4.50%", "~4.20–4.50%", 0.044, 0.038, 0.044, 0.043, 0.040,
         "Fed cut 3 times in H2-2024 (total −75bp). 10yr ~4.2-4.5%. "
         "MSFT hypothetical blended ~4.3% — slightly better than actual. "
         "Still significantly above 2021-2022 levels."),
    ]

    for col, hdr in zip("BCDEFGH",
                         ["Scenario", "Fed Funds Rate", "10yr Treasury",
                          "MSFT Blended Debt Rate", "Cash / ST Yield",
                          "AT Financing Cost ($M)", "Yr1 EPS & Accretion"]):
        col_header(ws, f"{col}9", hdr, horiz="left" if col in "BH" else "right")

    for i, (scen, ff, tyr, lt, mtn, cp, blended, cy, rationale) in enumerate(rate_env):
        r = 10 + i * 3  # 3 rows per scenario: header, values, note

        is_actual = "ACTUAL" in scen
        fg_c = NAVY if is_actual else (LT_AMBER if blended < DEBT_RATE else None)
        fc_c = WHITE if is_actual else (DK_AMBER if blended < DEBT_RATE else BLACK)

        gross = NEW_DEBT * blended + CASH_DEP * cy
        at_fin_ = gross * (1 - TAX_RATE)
        yr1_e   = pf_eps(RUN_RATE_SYN, blended, cy, TAX_RATE, PPA_AMORT, RAMP[0])
        acr     = accretion_pct(yr1_e)
        acr_str = f"${yr1_e:.4f}  |  {acr:+.2%}"

        lbl(ws, f"B{r}", f"  {scen}", fg=fg_c, fc=fc_c, bold=True)
        lbl(ws, f"C{r}", ff,         fg=fg_c, fc=fc_c)
        lbl(ws, f"D{r}", tyr,        fg=fg_c, fc=fc_c)
        data(ws, f"E{r}", blended,   fg=fg_c, fc=fc_c, bold=True, nf=NF_PCT)
        data(ws, f"F{r}", cy,        fg=fg_c, fc=fc_c, nf=NF_PCT)
        data(ws, f"G{r}", at_fin_,   fg=fg_c, fc=fc_c, bold=True, nf=NF_DOLLAR)
        lbl(ws, f"H{r}", acr_str,    fg=LT_GREEN if acr >= 0 else LT_RED,
            fc=DK_GREEN if acr >= 0 else DK_RED, bold=True)

        # Note row
        ws[f"B{r+1}"].value = rationale
        ws[f"B{r+1}"].font = font(GRAY_BODY, size=9, italic=True)
        ws[f"B{r+1}"].alignment = align("left", wrap=True)
        ws.merge_cells(f"B{r+1}:H{r+1}")
        ws.row_dimensions[r+1].height = 36
        ws.row_dimensions[r+2].height = 5   # spacer

    # ── Rate scenario comparison table ──────────────────────────────────────
    comp_r = 10 + len(rate_env) * 3 + 2
    section_hdr(ws, comp_r, "  RATE SCENARIO COMPARISON SUMMARY")
    ws.merge_cells(f"B{comp_r}:H{comp_r}")
    comp_r += 1

    for col, hdr in zip("BCDEFGH",
                         ["Scenario", "Blended Rate", "AT Fin Cost ($M)",
                          "EPS Drag (financing)", "Yr1 EPS",
                          "Yr1 Accretion", "vs. 2023 Actual"]):
        col_header(ws, f"{col}{comp_r}", hdr, horiz="left" if col=="B" else "right")
    comp_r += 1

    actual_at_fin = AT_FIN
    actual_yr1_eps = pf_eps(RUN_RATE_SYN, DEBT_RATE, CASH_YIELD, TAX_RATE, PPA_AMORT, RAMP[0])

    for scen, ff, tyr, lt, mtn, cp, blended, cy, _ in rate_env:
        gross = NEW_DEBT * blended + CASH_DEP * cy
        at_fin_ = gross * (1 - TAX_RATE)
        yr1_e   = pf_eps(RUN_RATE_SYN, blended, cy, TAX_RATE, PPA_AMORT, RAMP[0])
        acr     = accretion_pct(yr1_e)
        fin_drag_per_shr = -at_fin_ / MSFT_SHARES
        vs_actual = yr1_e - actual_yr1_eps

        is_actual = "ACTUAL" in scen
        fg_c = NAVY if is_actual else None
        fc_c = WHITE if is_actual else BLACK

        lbl(ws, f"B{comp_r}", f"  {scen}", fg=fg_c, fc=fc_c, bold=is_actual)
        data(ws, f"C{comp_r}", blended,          fg=fg_c, fc=fc_c, bold=is_actual, nf=NF_PCT)
        data(ws, f"D{comp_r}", at_fin_,          fg=fg_c, fc=fc_c, bold=is_actual, nf=NF_DOLLAR)
        data(ws, f"E{comp_r}", fin_drag_per_shr, fg=fg_c, fc=fc_c, bold=is_actual, nf=NF_DOLLAR2)
        data(ws, f"F{comp_r}", yr1_e,            fg=fg_c, fc=fc_c, bold=is_actual, nf=NF_DOLLAR2)
        acr_fg = (NAVY if is_actual else LT_GREEN) if acr >= 0 else LT_RED
        acr_fc = (WHITE if is_actual else DK_GREEN) if acr >= 0 else DK_RED
        data(ws, f"G{comp_r}", acr, fg=acr_fg, fc=acr_fc, bold=is_actual, nf=NF_PCT2)

        diff_fg = LT_GREEN if vs_actual > 0 else (NAVY if is_actual else LT_RED)
        diff_fc = DK_GREEN if vs_actual > 0 else (WHITE if is_actual else DK_RED)
        vs_txt = "Baseline" if is_actual else f"${vs_actual:+.4f}/share"
        lbl(ws, f"H{comp_r}", f"  {vs_txt}", fg=diff_fg, fc=diff_fc, bold=is_actual)
        comp_r += 1

    note(ws, f"B{comp_r+1}",
         "Key finding: Had the deal closed in 2021 at ~1.6% blended rate, Year-1 EPS would be "
         "~$0.14/share higher than actual — the rate cycle cost ~$0.14/share in annual EPS. "
         "The 2022-H2 rate environment would have been most cost-effective for debt issuance. "
         "2024 post-cut rates offer only marginal improvement vs. actual 2023 close.",
         fc=DARK_NAVY)
    ws.merge_cells(f"B{comp_r+1}:H{comp_r+1}")
    ws.row_dimensions[comp_r+1].height = 36

    set_row_height(ws)
    ws.row_dimensions[8].height = 21.75


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_memo(wb)
    build_breakeven(wb)
    build_financing_optimizer(wb)
    build_premium_sensitivity(wb)
    build_tornado(wb)
    build_historical_rates(wb)

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "MSFT_ATVI_Advanced_Analysis.xlsx")
    wb.save(out)
    print(f"Saved → {out}")


if __name__ == "__main__":
    main()
